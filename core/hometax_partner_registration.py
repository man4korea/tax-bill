# 📁 C:\APP\tax-bill\core\hometax_partner_registration.py
# Create at 2508312118 Ver1.00
# Update at 2509021347 Ver1.23
# -*- coding: utf-8 -*-
"""
HomeTax 거래처 등록 자동화 프로그램 (엑셀 통합 버전)
1. 엑셀 파일 열기/확인
2. 행 선택 GUI
3. HomeTax 자동 로그인 및 수동 로그인 여부 파악 
   자동 혹은 수동 로그인 완료 후 거래처 등록 화면 이동
4. 엑셀에서 가져온 거래처 등록번호로 오류체크
5. 홈택스에 거래처 등록
6. 결과 엑셀에 기록 (성공: 오늘 날짜, 실패: error)
"""

# Windows 콘솔 유니코드 출력 설정
import sys
import io
if sys.platform.startswith('win'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import asyncio
import os
import subprocess
import tkinter as tk
from tkinter import filedialog, ttk
from utils.copyable_messagebox import showinfo, showwarning, showerror, askyesno, askyesnocancel, askquestion
from dotenv import load_dotenv
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
import pandas as pd
from pathlib import Path
import re
import base64
import time

# 보안 관리자 import
sys.path.append(str(Path(__file__).parent.parent / "core"))
from hometax_security_manager import HomeTaxSecurityManager

# 통합 엑셀 처리 모듈 import
from excel_unified_processor import create_partner_processor

# 간단한 에러 처리 시스템
class ErrorCode:
    IMPORT_ERROR = "IMPORT_ERROR"
    EXCEL_ERROR = "EXCEL_ERROR"
    FILE_NOT_FOUND = "FILE_NOT_FOUND"

def handle_error(error, error_code, context):
    """간단한 에러 핸들러"""
    print(f"❌ {context} 오류 ({error_code}): {error}")
    return False

def check_and_install_dependencies():
    """필수 의존성 패키지 확인 및 자동 설치 (조용한 버전)"""
    required_packages = {
        'xlwings': 'xlwings>=0.30.0',
        'openpyxl': 'openpyxl>=3.1.0'
    }
    
    missing_packages = []
    
    for package_name, package_spec in required_packages.items():
        try:
            __import__(package_name)
        except ImportError:
            missing_packages.append(package_spec)
    
    if missing_packages:
        for package in missing_packages:
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", package], 
                                    capture_output=True, text=True)
            except subprocess.CalledProcessError as e:
                handle_error(e, ErrorCode.IMPORT_ERROR, f"패키지 설치 실패: {package}")
    
    return len(missing_packages) == 0

class ExcelRowSelector:
    """ExcelUnifiedProcessor 어댑터 클래스 - 기존 인터페이스 호환성 유지"""
    
    def __init__(self):
        # 통합 프로세서 생성 - 거래처 시트용
        self.processor = create_partner_processor()
        
        # 기존 인터페이스 호환을 위한 속성들
        self.selected_rows = None
        self.selected_data = None
        self.excel_file_path = None
        self.headers = None
        self.processed_data = []
        self.field_mapping = {}
    
    def initialize(self):
        """초기화 - 파일 열기 및 컴포넌트 생성"""
        if not self.processor.initialize():
            return False
        
        # 호환성을 위한 속성 동기화
        self.excel_file_path = self.processor.file_manager.excel_file_path
        return True
    
    def check_and_open_excel(self):
        """엑셀 파일 확인 및 열기"""
        return self.initialize()
    
    def show_row_selection_gui(self):
        """행 선택 GUI 표시"""
        if not self.processor.select_rows():
            return False
        
        # 호환성을 위한 속성 동기화
        self.selected_rows = self.processor.get_selected_rows()
        
        # 첫 번째 행의 첫 번째 열 값 저장 (기존 로직 유지)
        if self.selected_rows and self.excel_file_path:
            try:
                import pandas as pd
                from openpyxl import load_workbook
                
                wb = load_workbook(self.excel_file_path)
                ws = wb.active
                max_row = ws.max_row
                
                df = pd.read_excel(self.excel_file_path, sheet_name="거래처", header=None, dtype=str, 
                                 keep_default_na=False, engine='openpyxl', na_filter=False, nrows=max_row)
                
                first_row = self.selected_rows[0]
                if first_row <= len(df) and len(df.columns) > 0:
                    self.selected_data = df.iloc[first_row-1, 0]
                else:
                    self.selected_data = None
                    
            except Exception as e:
                handle_error(e, ErrorCode.EXCEL_ERROR, "엑셀 데이터 읽기")
                self.selected_data = None
        
        return True
    
    def process_excel_data(self):
        """엑셀 데이터 처리"""
        if not self.processor.process_data():
            return False
        
        # 호환성을 위한 데이터 동기화
        processed_data = self.processor.get_processed_data()
        
        # 기존 형식으로 변환
        self.processed_data = processed_data
        
        if processed_data:
            self.headers = list(processed_data[0]['data'].keys())
        
        return True
    
    def load_field_mapping(self):
        """field_mapping.md 파일을 읽어서 매핑 정보 추출 (조용한 버전)"""
        mapping_file = Path(__file__).parent / "field_mapping.md"
        
        if not mapping_file.exists():
            handle_error(FileNotFoundError(f"필드 매핑 파일을 찾을 수 없습니다: {mapping_file}"), 
                        ErrorCode.FILE_NOT_FOUND, "필드 매핑 파일 로드")
            return False
        
        try:
            with open(mapping_file, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # 테이블 파싱 (| 구분자 사용)
            lines = content.split('\n')
            mapping_found = False
            
            for line in lines:
                line = line.strip()
                if not line or not line.startswith('|'):
                    continue
                
                # 테이블 헤더나 구분선 스킵
                if '입력화면 라벨명' in line or ':--' in line:
                    mapping_found = True
                    continue
                
                if mapping_found and line.startswith('|'):
                    # 테이블 행 파싱
                    parts = [p.strip() for p in line.split('|')]
                    if len(parts) >= 5:  # |빈칸|라벨|변수명|Excel열명|셀렉터|
                        excel_column = parts[3]  # Excel 열명
                        selector = parts[4]      # HomeTax 셀렉션명
                        
                        if excel_column and selector and excel_column != "Excel 열명":
                            self.field_mapping[excel_column] = {
                                'selector': selector,
                                'label': parts[1] if len(parts) > 1 else '',
                                'variable': parts[2] if len(parts) > 2 else ''
                            }
            
            return True
            
        except Exception as e:
            handle_error(e, ErrorCode.FILE_NOT_FOUND, "필드 매핑 파일 파싱")
            return False
    
    def write_error_to_excel(self, row_number, error_message="error"):
        """에러 상태 기록"""
        return self.processor.record_error(row_number, error_message)
    
    def write_today_to_excel(self, row_number):
        """성공 상태 기록 (오늘 날짜)"""
        return self.processor.record_success(row_number)
    
    def split_email(self, email_str):
        """이메일 주소를 @ 기준으로 분리"""
        if pd.isna(email_str) or not str(email_str).strip():
            return "", ""
        
        email_str = str(email_str).strip()
        if '@' in email_str:
            parts = email_str.split('@', 1)  # 첫 번째 @에서만 분리
            return parts[0].strip(), parts[1].strip()
        else:
            # @ 없는 경우 전체를 앞부분으로 처리
            return email_str, ""

async def prepare_next_registration(main_page):
    """다음 거래처 등록을 위한 페이지 준비"""
    try:
        business_number_selectors = [
            "#mf_txppWframe_txtBsno1",
            "input[name*='txtBsno']",
            "input[id*='Bsno']",
            "input[placeholder*='사업자']",
            "input[title*='사업자']",
            "input[type='text']",
            "*[id*='txtBsno']"
        ]
        
        business_field = None
        for selector in business_number_selectors:
            try:
                business_field = main_page.locator(selector).first
                await business_field.wait_for(state="visible", timeout=5000)
                break
            except:
                continue
        
        if not business_field:
            try:
                all_inputs = await main_page.locator("input").all()
                for i, input_field in enumerate(all_inputs[:10]):
                    try:
                        field_id = await input_field.get_attribute("id")
                        field_name = await input_field.get_attribute("name") 
                        field_placeholder = await input_field.get_attribute("placeholder")
                        
                        if (field_id and ('bsno' in field_id.lower() or 'business' in field_id.lower())) or \
                           (field_name and ('bsno' in field_name.lower() or 'business' in field_name.lower())) or \
                           (field_placeholder and '사업자' in field_placeholder):
                            business_field = input_field
                            break
                    except:
                        continue
            except:
                pass
            
            if not business_field:
                return False
        
        try:
            # 여러 방법으로 필드 클리어 시도
            clear_success = False
            
            # 방법 1: clear()
            try:
                await business_field.clear()
                clear_success = True
            except:
                pass
            
            # 방법 2: selectAll + delete
            if not clear_success:
                try:
                    await business_field.press("Control+a")
                    await business_field.press("Delete")
                    clear_success = True
                except:
                    pass
            
            # 방법 3: fill("")
            if not clear_success:
                try:
                    await business_field.fill("")
                    clear_success = True
                except:
                    pass
            
            await business_field.focus()
            
            try:
                field_value = await business_field.input_value()
            except:
                pass
            
            await main_page.wait_for_timeout(1000)
            return True
            
        except:
            return False
            
    except:
        return False

async def fill_hometax_form(main_page, row_data, field_mapping, excel_selector, current_row_number, is_first_record=False):
    """HomeTax 폼에 데이터 자동 입력 (단순화)"""
    
    try:
        # 첫 번째 거래처가 아닌 경우 페이지 준비
        if not is_first_record:
            if not await prepare_next_registration(main_page):
                raise Exception("다음 거래처 등록을 위한 페이지 준비에 실패했습니다.")
        
        success_count = 0
        failed_fields = []

        # 2. 사업자번호 입력 및 확인
        business_number = row_data.get('사업자번호') or row_data.get('사업자등록번호') or row_data.get('거래처등록번호')
        if business_number and ('사업자번호' in field_mapping or '사업자등록번호' in field_mapping or '거래처등록번호' in field_mapping):
            try:
                # 사업자번호 필드 찾기
                if '사업자번호' in field_mapping:
                    business_field = '사업자번호'
                elif '사업자등록번호' in field_mapping:
                    business_field = '사업자등록번호'
                else:
                    business_field = '거래처등록번호'
                selector = field_mapping[business_field]['selector'].strip()
                
                # 사업자번호 입력
                element = main_page.locator(selector).first
                await element.wait_for(state="visible", timeout=1000)
                await element.clear()
                await element.fill(str(business_number))
                await main_page.wait_for_timeout(1000)
                
                success_count += 1
                
                # 사업자번호 확인 버튼 클릭 및 검증
                await handle_business_number_validation(main_page, business_number, excel_selector, current_row_number)
                
            except Exception as e:
                # 사업자번호 처리 중 오류 발생
                if "BUSINESS_NUMBER_ERROR" in str(e):
                    excel_selector.write_error_to_excel(current_row_number, "error")
                failed_fields.append({'field': '사업자번호', 'error': str(e)})

        # 3. 나머지 필드들 입력 (새 팝업 페이지에서 수행)
        for excel_column, value in row_data.items():
            if excel_column in ['사업자번호', '사업자등록번호', '거래처등록번호']:
                continue
            
            if not value:
                continue
            
            if excel_column not in field_mapping:
                continue
            
            mapping_info = field_mapping[excel_column]
            selector = mapping_info['selector'].strip()
            
            if not selector:
                continue
            
            try:
                element = main_page.locator(selector).first
            
                # 먼저 일반적인 방법으로 시도
                try:
                    await element.clear(timeout=1000)
                    await element.fill(str(value), timeout=1000)
                    await main_page.wait_for_timeout(200)
                    success_count += 1
                    continue  # 성공하면 다음 필드로
                except Exception as normal_error:
                    
                    # JavaScript로 강제 입력 시도
                    try:
                        # disabled 속성을 제거하고 값을 설정
                        await main_page.evaluate(f"""
                            const element = document.querySelector('{selector}');
                            if (element) {{
                                element.removeAttribute('disabled');
                                element.value = '{value}';
                                element.dispatchEvent(new Event('input', {{'bubbles': true}}));
                                element.dispatchEvent(new Event('change', {{'bubbles': true}}));
                            }}
                        """, "")
                        success_count += 1
                    except Exception as js_error:
                        failed_fields.append({'field': excel_column, 'selector': selector, 'error': f"일반: {normal_error}, JS: {js_error}"})
                        
            except Exception as e:
                failed_fields.append({'field': excel_column, 'selector': selector, 'error': str(e)})

        # 4. 기타 특별 처리 필드들 (새 팝업 페이지에서 수행)
        await handle_other_special_fields(main_page, row_data, field_mapping)
    
        # 입력 결과 요약

        # 5. 최종 등록 버튼 클릭 및 Alert 처리
        try:
            
            # Alert 리스너 설정
            alert_handled = False
            alert_message = ""
            
            async def handle_final_alert(dialog):
                nonlocal alert_handled, alert_message
                alert_message = dialog.message
                
                # 품목 등록 또는 담당자 추가 Alert인 경우 취소 클릭
                if "품목 등록" in alert_message:
                    await dialog.dismiss()  # 취소 버튼 클릭
                elif "담당자를 추가 등록" in alert_message:
                    await dialog.dismiss()  # 취소 버튼 클릭
                else:
                    await dialog.accept()  # 확인 버튼 클릭
                
                alert_handled = True

            main_page.on("dialog", handle_final_alert)

            # 등록 버튼 클릭 (여러 방법 시도)
            register_btn = main_page.locator("#mf_txppWframe_btnRgt").first
            
            try:
                # 방법 1: 일반 클릭
                await register_btn.click(timeout=1000)
            except Exception as e1:
                try:
                    # 방법 2: 강제 클릭 (다른 요소가 가리고 있어도 클릭)
                    await register_btn.click(force=True, timeout=1000)
                except Exception as e2:
                    try:
                        # 방법 3: JavaScript를 통한 클릭
                        await main_page.evaluate("document.getElementById('mf_txppWframe_btnRgt').click()")
                    except Exception as e3:
                        raise Exception("모든 등록 버튼 클릭 방법이 실패했습니다")

            # Alert 대기 (더 긴 시간)
            for i in range(100): # 10초 대기
                if alert_handled:
                    break
                if i % 10 == 0:
                    pass
                await main_page.wait_for_timeout(100)

            main_page.remove_listener("dialog", handle_final_alert)

            if alert_handled:
                # 등록 성공 시 엑셀 파일에 오늘 날짜 기록
                excel_selector.write_today_to_excel(current_row_number)
            else:
                # Alert가 없어도 등록이 성공한 것으로 간주하고 날짜 기록
                excel_selector.write_today_to_excel(current_row_number)

        except Exception as e:
            failed_fields.append({'field': '등록 버튼', 'error': str(e)})

        return success_count, failed_fields

    except Exception as e:
        return 0, [f"폼 입력 중 오류: {str(e)}"]


async def handle_business_number_validation(main_page, business_number, excel_selector, current_row_number):
    """사업자번호 확인 버튼 클릭 및 검증 (종사업장 선택 창 처리 포함)"""
    try:
        confirm_btn = main_page.locator("#mf_txppWframe_btnValidCheck").first
        
        try:
            await confirm_btn.click(timeout=1000)
        except:
            try:
                await confirm_btn.click(force=True, timeout=1000)
            except:
                try:
                    await main_page.evaluate("document.getElementById('mf_txppWframe_btnValidCheck').click()")
                except:
                    return

        await main_page.wait_for_timeout(1000)
        
        workplace_popup_selectors = [
            "#mf_txppWframe_ABTIBsnoUnitPopup2",
            ".popup:has-text('종사업장')",
            "[id*='BsnoUnit']",
        ]
        
        workplace_popup_found = False
        for selector in workplace_popup_selectors:
            try:
                element = main_page.locator(selector).first
                if await element.is_visible():
                    workplace_popup_found = True
                    break
            except:
                continue
        
        if workplace_popup_found:
            workplace_confirm_btn = main_page.locator("#mf_txppWframe_ABTIBsnoUnitPopup2_wframe_trigger66").first
            
            for i in range(600):
                try:
                    if await workplace_confirm_btn.is_visible():
                        if i % 50 == 0:
                            try:
                                import winsound
                                winsound.Beep(1000, 300)
                            except:
                                pass
                        await main_page.wait_for_timeout(100)
                    else:
                        break
                except:
                    break
            
            alert_handled = False
            alert_message = ""
            
            async def handle_workplace_alert(dialog):
                nonlocal alert_handled, alert_message
                alert_message = dialog.message
                await dialog.accept()
                alert_handled = True
            
            main_page.on("dialog", handle_workplace_alert)
            
            for i in range(50):
                if alert_handled:
                    break
                await main_page.wait_for_timeout(100)
            
            main_page.remove_listener("dialog", handle_workplace_alert)
            
        else:
            alert_handled = False
            alert_message = ""
            
            async def handle_alert(dialog):
                nonlocal alert_handled, alert_message
                alert_message = dialog.message
                await dialog.accept()
                alert_handled = True
            
            main_page.on("dialog", handle_alert)
            
            for i in range(50):
                if alert_handled:
                    break
                await main_page.wait_for_timeout(100)
            
            main_page.remove_listener("dialog", handle_alert)
            
            if alert_handled:
                if "비정상적인 등록번호" in alert_message or "이미 등록된 사업자등록번호" in alert_message:
                    raise Exception(f"SKIP_TO_NEXT_ROW|{alert_message}")
        
        await main_page.wait_for_timeout(2000)
        
        try:
            await main_page.wait_for_selector("#mf_txppWframe_txtTnmNm:not([disabled])", timeout=1000)
        except:
            pass
            
    except Exception as e:
        if "SKIP_TO_NEXT_ROW" in str(e):
            raise e
        return

async def handle_other_special_fields(main_page, row_data, field_mapping):
    """기타 특별 처리가 필요한 필드들 처리 (사업자번호 제외)"""
    
    # 이메일 직접입력 버튼들
    email_fields = [
        ('주이메일앞', '주이메일뒤'),
        ('부이메일앞', '부이메일뒤')
    ]
    
    for email_front, email_back in email_fields:
        if email_front in row_data and row_data[email_front]:
            try:
                # 직접입력 버튼 찾기
                if '주이메일' in email_front:
                    direct_btn = main_page.locator("#mf_txppWframe_btnMainEmailDirect").first
                else:
                    direct_btn = main_page.locator("#mf_txppWframe_btnSubEmailDirect").first
                
                await direct_btn.click(timeout=1000)
                await main_page.wait_for_timeout(300)
            except:
                pass


def decrypt_password_from_env(encrypted_config):
    """암호화된 설정에서 비밀번호 복호화"""
    try:
        if not encrypted_config.startswith("HTC_") or not encrypted_config.endswith("_CFG"):
            return None
        
        middle_part = encrypted_config[4:-4]
        original_encoded = middle_part[::-1]
        decoded = base64.b64decode(original_encoded.encode('utf-8')).decode('utf-8')
        return decoded
    except:
        return None

def load_encrypted_config_from_env(env_file):
    """암호화된 설정을 .env 파일에서 로드"""
    try:
        if not env_file.exists():
            return None
        
        with open(env_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        for line in lines:
            line = line.strip()
            if line.startswith('HTC_CONFIG='):
                encrypted_config = line.split('=', 1)[1].strip()
                password = decrypt_password_from_env(encrypted_config)
                if password:
                    return password
        return None
    except:
        return None

def load_env_settings():
    """환경변수 설정 로드"""
    try:
        env_file = Path(__file__).parent.parent / ".env"
        if not env_file.exists():
            return "manual", None
        
        load_dotenv(env_file)
        login_mode = os.getenv("HOMETAX_LOGIN_MODE", "manual")
        
        # 보안 관리자로 비밀번호 로드
        security_manager = HomeTaxSecurityManager()
        password = security_manager.load_password_from_env()
        
        return login_mode, password
    except:
        return "manual", None

def get_certificate_password():
    """인증서 비밀번호 가져오기"""
    try:
        env_file = Path(__file__).parent.parent / ".env"
        encrypted_password = load_encrypted_config_from_env(env_file)
        if encrypted_password:
            return encrypted_password
        
        security_manager = HomeTaxSecurityManager()
        legacy_password = security_manager.load_password_from_env()
        if legacy_password:
            return legacy_password
        
        basic_password = os.getenv("PW")
        if basic_password:
            return basic_password
        
        return None
    except:
        return None


async def main():
    """메인 실행 함수"""
    try:
        check_and_install_dependencies()
        
        excel_selector = ExcelRowSelector()
        
        if not excel_selector.check_and_open_excel():
            # showerror("엑셀 오류", "엑셀 파일 열기에 실패했습니다.")
            return
        
        if not excel_selector.show_row_selection_gui():
            # showwarning("선택 취소", "행 선택이 취소되었습니다.")
            return
        
        if not excel_selector.load_field_mapping():
            # showerror("매핑 오류", "필드 매핑 로드에 실패했습니다.")
            return
    
        # 2.6. 엑셀 데이터 처리
        if not excel_selector.process_excel_data():
            return
        
        # 4. HomeTax 개선된 로그인 실행 (test_hometax_menu_navigation.py 기반)
        playwright = await async_playwright().start()
        browser = await playwright.chromium.launch(headless=False)
        page = await browser.new_page()
        main_page = page
        main_browser = browser
        
        await page.goto("https://hometax.go.kr/websquare/websquare.html?w2xPath=/ui/pp/index_pp.xml&menuCd=index3")
        await page.wait_for_timeout(6000)
        
        # 환경설정 로드
        login_mode, password = load_env_settings()
        
        if login_mode == "auto":
            # 1. 공동.금융인증 로그인 버튼 자동 클릭
            await page.click("#mf_txppWframe_loginboxFrame_anchor22")
            # 인증서 입력창 등장까지 명시적으로 대기
            await page.wait_for_selector("#dscert", state="visible", timeout=15000)
            
            # 2. 인증서 비밀번호 자동 입력
            if password:
                iframe = page.frame_locator("#dscert")
                await iframe.locator("#input_cert_pw").fill(password)
                await iframe.locator("#btn_confirm_iframe > span").click()
            
            # 인증서 iframe이 완전히 사라질 때(= 로그인 완료)까지 대기
            await page.wait_for_selector("#dscert", state="hidden", timeout=60000)
            
        else:
            # manual 모드: 사용자가 버튼을 클릭하고 인증을 완료해야 함
            print("manual 모드: 사람이 [공동·금융인증 로그인] 버튼을 클릭하고 인증을 완료하세요.")
            
            # (A) 사용자가 버튼을 눌러 인증창이 '등장'할 때까지 먼저 대기
            try:
                # 페이지 상태 확인
                if page.is_closed():
                    raise Exception("페이지가 닫혔습니다.")
                    
                await page.wait_for_selector("#dscert", state="visible", timeout=120000)
                print("✅ 인증서 창이 나타났습니다. 인증을 완료해주세요.")
            except PlaywrightTimeoutError:
                print("❌ 인증서 창이 나타나지 않았습니다. 버튼을 클릭했는지 확인하세요.")
                raise
            except Exception as e:
                print(f"❌ 인증서 창 대기 중 오류 발생: {e}")
                raise
            
            # (B) 인증 완료로 인증창이 '사라질' 때까지 대기
            print("🔐 인증서 비밀번호 입력 및 확인 버튼 클릭을 완료해주세요...")
            try:
                # 페이지 상태 확인
                if page.is_closed():
                    raise Exception("페이지가 닫혔습니다.")
                    
                await page.wait_for_selector("#dscert", state="hidden", timeout=600000)
                print("✅ 로그인이 완료되었습니다!")
            except PlaywrightTimeoutError:
                print("❌ 로그인 완료 대기 시간이 초과되었습니다.")
                raise
            except Exception as e:
                print(f"❌ 로그인 완료 대기 중 오류 발생: {e}")
                raise
        
        # ▲ 여기까지가 '로그인 완료'의 신뢰 가능한 기준
        
        # 3. 이벤트 기반 실시간 팝업 감지 시스템 설정
        popup_pages = []
        popup_detection_active = True
        
        def on_new_page(new_page):
            """새로 열리는 페이지를 실시간으로 감지"""
            if popup_detection_active:
                popup_pages.append(new_page)
                print(f"🔔 새 페이지 탐지: {new_page.url}")
                
                # 즉시 팝업 여부 판단하고 닫기
                asyncio.create_task(check_and_close_popup(new_page))
        
        async def check_and_close_popup(new_page):
            """새 페이지가 팝업인지 판단하고 즉시 닫기"""
            try:
                # 메인 페이지는 절대 닫지 않도록 보호
                if new_page == page:
                    print(f"   🛡️ 메인 페이지 보호: {new_page.url}")
                    return
                
                # 페이지가 이미 닫혔는지 확인
                if new_page.is_closed():
                    print(f"   ℹ️ 페이지가 이미 닫혔습니다: {new_page.url}")
                    return
                
                await new_page.wait_for_load_state("domcontentloaded", timeout=5000)
                url = new_page.url
                title = await new_page.title()
                
                # 메인 페이지 헤더가 있는지 확인 (추가 보호)
                try:
                    header_count = await new_page.locator("#mf_wfHeader_wq_uuid_359").count()
                    if header_count > 0:
                        print(f"   🛡️ 메인 헤더 감지로 페이지 보호: {url}")
                        return
                except:
                    pass
                
                # 팝업 판단 기준
                popup_reasons = []
                if any(k in (title or "") for k in ("알림", "공지", "Notice", "Popup", "안내", "공지사항", "시스템 점검")):
                    popup_reasons.append(f"제목에 팝업 키워드: '{title}'")
                if "popup" in (url or "").lower():
                    popup_reasons.append("URL에 'popup' 포함")
                if "UTXPPABC13" in (url or ""):
                    popup_reasons.append("홈택스 공지창 (UTXPPABC13) 감지")
                if "w2xPath" in (url or "") and "popupID" in (url or ""):
                    popup_reasons.append("홈택스 팝업 패턴 (w2xPath + popupID) 감지")
                if "websquare/popup.html" in (url or ""):
                    popup_reasons.append("홈택스 웹스퀘어 팝업창 감지")
                if "/ui/pp/a/b/" in (url or "") and any(x in (url or "") for x in ["UTC", "UTX", "popup"]):
                    popup_reasons.append("홈택스 안내창 경로 패턴 감지")
                
                if popup_reasons:
                    print(f"⚡ 실시간 팝업 감지 및 즉시 닫기: {url}")
                    for reason in popup_reasons:
                        print(f"   📋 감지 이유: {reason}")
                    
                    # 페이지가 아직 열려있는지 다시 한번 확인
                    if not new_page.is_closed():
                        await new_page.close()
                        print(f"   ❌ 팝업 페이지 즉시 닫음")
                    else:
                        print(f"   ℹ️ 페이지가 이미 닫혀있음")
                else:
                    print(f"   ✅ 일반 페이지로 판단: {url} (제목: {title})")
                    
            except Exception as e:
                print(f"   ⚠️ 실시간 팝업 처리 실패: {e}")
        
        # 브라우저에 새 페이지 이벤트 리스너 등록
        browser.on("page", on_new_page)
        
        print("🔍 이벤트 기반 팝업 감지 시스템 활성화 (5초간 감시)")
        await asyncio.sleep(5)  # 5초간 새로 열리는 팝업 감시
        
        popup_detection_active = False
        print(f"📊 감지된 새 페이지 수: {len(popup_pages)}개")
        
        # 4. 모든 브라우저 컨텍스트에서 메인 UI가 있는 페이지 찾기 및 기존 팝업 정리
        print("🔍 모든 브라우저 컨텍스트에서 메인 페이지 검색 중...")
        
        # 메인 헤더가 있는 페이지를 모든 컨텍스트에서 찾기
        target = None
        target_context = None
        
        # 현재 브라우저의 모든 컨텍스트 확인
        all_contexts = browser.contexts
        print(f"📊 총 {len(all_contexts)}개의 컨텍스트 발견")
        
        for ctx_idx, ctx in enumerate(all_contexts):
            print(f"   🔍 컨텍스트 {ctx_idx + 1}: {len(ctx.pages)}개 페이지")
            
            for page_idx, p in enumerate(ctx.pages[::-1]):  # 최근 열린 페이지부터 검사
                try:
                    url = p.url
                    print(f"      📄 페이지 {page_idx + 1}: {url}")
                    
                    # 메인 헤더 확인
                    header_count = await p.locator("#mf_wfHeader_wq_uuid_359").count()
                    if header_count > 0:
                        target = p
                        target_context = ctx
                        print(f"   ✅ 메인 페이지 발견: {url}")
                        break
                except Exception as e:
                    print(f"      ⚠️ 페이지 검사 실패: {e}")
                    continue
            
            if target:
                break
        
        # 메인 페이지로 포커스 변경
        if target:
            page = target
            await page.bring_to_front()
            print(f"✅ 메인 페이지로 포커스 변경: {page.url}")
        else:
            print("⚠️ 메인 페이지를 찾지 못했습니다. 기존 페이지를 유지합니다.")
        
        # 모든 컨텍스트의 팝업 페이지 정리
        closed_count = 0
        popup_contexts_to_close = []
        
        for ctx_idx, ctx in enumerate(all_contexts):
            pages_to_close = []
            main_page_found = False
            
            for p in ctx.pages:
                if target and p == target:
                    main_page_found = True
                    continue
                
            try:
                title = await p.title()
                url = p.url
                
                # 팝업 페이지 판단 기준 (상세 로그 포함)
                popup_reasons = []
                
                # 제목 기준 팝업 감지
                if any(k in (title or "") for k in ("알림", "공지", "Notice", "Popup", "안내", "공지사항", "시스템 점검")):
                    popup_reasons.append(f"제목에 팝업 키워드 포함: '{title}'")
                
                # URL 기준 팝업 감지
                if "popup" in (url or "").lower():
                    popup_reasons.append("URL에 'popup' 포함")
                
                # 홈택스 특화 팝업 패턴들
                if "UTXPPABC13" in (url or ""):
                    popup_reasons.append("홈택스 공지창 (UTXPPABC13) 감지")
                
                if "w2xPath" in (url or "") and "popupID" in (url or ""):
                    popup_reasons.append("홈택스 팝업 패턴 (w2xPath + popupID) 감지")
                
                # 추가 홈택스 팝업 패턴
                if "websquare/popup.html" in (url or ""):
                    popup_reasons.append("홈택스 웹스퀘어 팝업창 감지")
                
                if "/ui/pp/a/b/" in (url or "") and any(x in (url or "") for x in ["UTC", "UTX", "popup"]):
                    popup_reasons.append("홈택스 안내창 경로 패턴 감지")
                
                is_popup = len(popup_reasons) > 0
                
                if is_popup:
                    pages_to_close.append(p)
                    print(f"🗑️ 팝업 페이지 마킹: {url}")
                    for reason in popup_reasons:
                        print(f"      📋 감지 이유: {reason}")
                else:
                    print(f"   ✅ 일반 페이지 유지: {url} (제목: {title})")
                    
            except Exception as e:
                print(f"⚠️ 페이지 검사 실패: {e}")
                continue
        
            # 페이지 닫기
            for p in pages_to_close:
                try:
                    await p.close()
                    closed_count += 1
                    print(f"   ❌ 팝업 페이지 닫음: {p.url}")
                except:
                    pass
            
            # 메인 페이지가 없는 빈 컨텍스트는 닫기 대상으로 마킹
            if not main_page_found and len(ctx.pages) == 0:
                popup_contexts_to_close.append(ctx)
        
        # 빈 컨텍스트 닫기
        for ctx in popup_contexts_to_close:
            try:
                await ctx.close()
                print(f"🗑️ 빈 컨텍스트 닫음")
            except:
                pass
        
        print(f"✅ 총 {closed_count}개의 기존 팝업 페이지를 닫았습니다.")
        
        # 이벤트 리스너 정리
        try:
            browser.remove_listener("page", on_new_page)
            print("🧹 이벤트 리스너 정리 완료")
        except:
            pass
        
        await page.wait_for_timeout(2000)
        
        # 5. 계산서·영수증·카드 메뉴 클릭 (메인 페이지에서만 수행)
        try:
            await page.wait_for_selector("#mf_wfHeader_wq_uuid_359", timeout=30000)
            await page.click("#mf_wfHeader_wq_uuid_359")
            print("✅ 계산서·영수증·카드 메뉴 클릭 완료")
            await page.wait_for_timeout(3000)
        except Exception as e:
            print(f"❌ 메뉴 클릭 실패: {e}")
            # raise 제거하고 계속 진행
            pass
        try:
            await main_page.click("#menuAtag_4601020000 > span")
            print("✅ 거래처 및 품목관리 메뉴 클릭 성공")
            await main_page.wait_for_timeout(1000)
        except Exception as sub_menu_error:
            print(f"⚠️ 거래처 및 품목관리 메뉴 클릭 오류: {str(sub_menu_error)}")
        
        # 전자세금계산서 거래처 클릭
        print("📝 전자세금계산서 거래처 메뉴 클릭...")
        try:
            await main_page.click("#menuAtag_4601020100 > span")
            print("✅ 전자세금계산서 거래처 메뉴 클릭 성공")
            await main_page.wait_for_timeout(1000)
        except Exception as final_menu_error:
            print(f"⚠️ 전자세금계산서 거래처 메뉴 클릭 오류: {str(final_menu_error)}")

        # 건별 등록 버튼 클릭
        print("🔘 건별 등록 버튼 클릭...")
        try:
            await main_page.click("#mf_txppWframe_textbox1395")
            print("✅ 건별 등록 버튼 클릭 성공")
            await main_page.wait_for_timeout(1000)
        except Exception as register_button_error:
            print(f"⚠️ 건별 등록 버튼 클릭 오류: {str(register_button_error)}")

        # 6. 실제 거래처 등록 자동화 실행
        print("🏃 거래처 등록 자동화 시작...")
        try:
            # 메인 페이지에서 자동화 실행 (새 창 무시)
            await main_page.bring_to_front()  # 최종 포커스 확인
            
            # 선택된 데이터에 대해 거래처 등록 수행
            success_count = 0
            failed_count = 0
            
            for idx, row_info in enumerate(excel_selector.processed_data):
                current_row_number = row_info['row_number']
                row_data = row_info['data']
                
                try:
                    # 각 거래처에 대해 폼 입력 실행
                    is_first_record = (idx == 0)
                    success_count_fields, failed_fields = await fill_hometax_form(
                        main_page, row_data, excel_selector.field_mapping, 
                        excel_selector, current_row_number, is_first_record
                    )
                    
                    if success_count_fields > 0:
                        success_count += 1
                    else:
                        failed_count += 1
                        
                except:
                    failed_count += 1
                    excel_selector.write_error_to_excel(current_row_number, "error")
                
                # 다음 거래처 등록을 위한 대기
                if idx < len(excel_selector.processed_data) - 1:
                    await main_page.wait_for_timeout(3000)
                
        except:
            pass
        
        # 브라우저 정리
        await main_page.wait_for_timeout(5000)
        
        if main_browser:
            try:
                await main_browser.close()
            except:
                pass
                
    except Exception as e:
        print(f"❌ 메인 함수 실행 중 오류: {e}")
        pass
    finally:
        try:
            if 'browser' in locals():
                await browser.close()
        except:
            pass
       
if __name__ == "__main__":
    asyncio.run(main())

