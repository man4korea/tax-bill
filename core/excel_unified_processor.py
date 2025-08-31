# 📁 C:\APP\tax-bill\core\excel_unified_processor.py
# Create at 2508312118 Ver1.00
# -*- coding: utf-8 -*-
"""
HomeTax 엑셀 데이터 통합 처리 모듈
거래처 시트와 거래명세표 시트의 공통 기능을 통합
"""

import os
import sys
import subprocess
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from pathlib import Path
import re
from datetime import datetime
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple, Any


@dataclass
class SheetConfig:
    """시트별 설정"""
    sheet_name: str          # 시트 이름
    target_filename: str     # 파일 이름  
    status_column: int       # 상태 기록할 컬럼 (1-based)
    business_number_column: str  # 사업자번호 컬럼명
    company_name_column: str     # 상호 컬럼명
    
    @classmethod
    def get_partner_config(cls):
        """거래처 시트 설정"""
        return cls(
            sheet_name="거래처",
            target_filename="세금계산서.xlsx", 
            status_column=1,  # A열
            business_number_column="사업자등록번호",
            company_name_column="상호"
        )
    
    @classmethod 
    def get_transaction_config(cls):
        """거래명세표 시트 설정"""
        return cls(
            sheet_name="거래명세표",
            target_filename="세금계산서.xlsx",
            status_column=17,  # Q열
            business_number_column="등록번호", 
            company_name_column="상호"
        )


class ExcelFileManager:
    """엑셀 파일 관리 클래스"""
    
    def __init__(self, config: SheetConfig):
        self.config = config
        self.excel_file_path = None
    
    def check_and_open_excel(self) -> bool:
        """엑셀 파일 확인 및 열기 (3단계 프로세스)"""
        target_file = self._get_target_file_path()
        
        print("=== 엑셀 파일 확인 (3단계 체크) ===")
        
        # 1단계: 이미 열려있는가?
        if self._check_already_opened():
            return True
            
        # 2단계: 문서 폴더에 파일이 있는가?
        if self._open_from_documents(target_file):
            return True
            
        # 3단계: 파일 선택 창
        return self._open_from_dialog(target_file)
    
    def _get_target_file_path(self) -> str:
        """대상 파일 경로 반환"""
        return os.path.expanduser(f"~/OneDrive/문서/{self.config.target_filename}")
    
    def _check_already_opened(self) -> bool:
        """이미 열린 파일 확인"""
        print(f"1단계: '{self.config.target_filename}' 파일이 이미 열려있는지 확인...")
        
        try:
            result = subprocess.run(['tasklist', '/fi', 'imagename eq excel.exe'], 
                                  capture_output=True, text=True)
            if 'excel.exe' not in result.stdout.lower():
                print("   Excel 프로세스가 실행되지 않았습니다.")
                return False
                
            # xlwings로 열린 파일 확인
            try:
                import xlwings as xw
                app = xw.apps.active if xw.apps else None
                
                if app and hasattr(app, 'books'):
                    for book in app.books:
                        if book.name.lower() == self.config.target_filename.lower():
                            print(f"   ✅ '{book.name}' 파일이 이미 열려있습니다!")
                            self.excel_file_path = book.fullname
                            return True
                            
            except ImportError:
                print("   xlwings가 설치되지 않았습니다.")
            except Exception as e:
                print(f"   xlwings 확인 중 오류: {e}")
                
        except Exception as e:
            print(f"   프로세스 확인 중 오류: {e}")
            
        return False
    
    def _open_from_documents(self, target_file: str) -> bool:
        """문서 폴더에서 파일 열기"""
        print(f"2단계: 문서 폴더에 '{self.config.target_filename}' 파일이 있는지 확인...")
        
        if os.path.exists(target_file):
            print(f"   ✅ 파일 발견: {target_file}")
            try:
                os.startfile(target_file)
                self.excel_file_path = target_file
                
                # Excel 로딩 대기 및 포커스 복원
                import time
                time.sleep(3)
                self._restore_console_focus()
                
                print(f"   ✅ '{self.config.target_filename}' 파일이 열렸습니다!")
                return True
                
            except Exception as e:
                print(f"   ❌ 파일 열기 실패: {e}")
        else:
            print(f"   ❌ 문서 폴더에 '{self.config.target_filename}' 파일이 없습니다.")
            
        return False
    
    def _open_from_dialog(self, target_file: str) -> bool:
        """파일 선택 창에서 파일 열기"""
        print(f"3단계: 파일 선택 창에서 '{self.config.target_filename}' 파일을 선택해주세요...")
        
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        
        messagebox.showinfo(
            "파일 선택", 
            f"다음 창에서 '{self.config.target_filename}' 파일을 선택해주세요."
        )
        
        file_path = filedialog.askopenfilename(
            title=f"'{self.config.target_filename}' 파일을 선택하세요",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=os.path.dirname(target_file) if os.path.exists(os.path.dirname(target_file)) else os.path.expanduser("~/Documents")
        )
        
        if file_path:
            print(f"   ✅ 선택된 파일: {file_path}")
            try:
                os.startfile(file_path)
                self.excel_file_path = file_path
                
                # Excel 로딩 대기 및 포커스 복원
                import time
                time.sleep(3)
                self._restore_console_focus()
                
                print(f"   ✅ 파일이 열렸습니다!")
                root.destroy()
                return True
                
            except Exception as e:
                print(f"   ❌ 파일 열기 실패: {e}")
                messagebox.showerror("오류", f"파일을 열 수 없습니다: {e}")
        else:
            print("   ❌ 파일이 선택되지 않았습니다.")
            messagebox.showerror("오류", "파일을 선택하지 않으면 프로그램을 계속할 수 없습니다.")
            
        root.destroy()
        return False
    
    def _restore_console_focus(self):
        """콘솔 포커스 복원"""
        try:
            import win32gui
            console_hwnd = win32gui.GetConsoleWindow()
            if console_hwnd:
                win32gui.SetForegroundWindow(console_hwnd)
                print("   ✅ 포커스를 콘솔로 복원")
        except:
            pass


class RowSelector:
    """행 선택 GUI 클래스"""
    
    def __init__(self, config: SheetConfig, file_path: str):
        self.config = config
        self.file_path = file_path
        self.selected_rows = None
    
    def show_row_selection_gui(self) -> bool:
        """행 선택 GUI 표시"""
        print("\n=== 행 선택 GUI ===")
        
        root = tk.Tk()
        root.title("행 선택")
        root.resizable(False, False)
        
        # 화면 중앙 위치
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        window_width = 500
        window_height = 550
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 4
        root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # UI 구성
        self._create_gui_components(root)
        
        root.mainloop()
        return self.selected_rows is not None
    
    def _create_gui_components(self, root):
        """GUI 컴포넌트 생성"""
        main_frame = ttk.Frame(root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 제목
        title_label = ttk.Label(main_frame, text=f"{self.config.sheet_name} 시트에서 처리할 행을 선택하세요", 
                               font=('맑은 고딕', 14, 'bold'))
        title_label.pack(pady=(0, 20))
        
        # 안내 메시지
        guide_frame = ttk.LabelFrame(main_frame, text="행 선택 방법", padding="10")
        guide_frame.pack(fill=tk.X, pady=(0, 20))
        
        guide_text = """• 단일 행: 2
• 복수 행: 2,4,8  
• 범위: 2-8
• 혼합: 2,5-7,10

예시: 2행, 5~7행, 10행을 처리하려면 → 2,5-7,10"""
        
        ttk.Label(guide_frame, text=guide_text, justify=tk.LEFT).pack(anchor=tk.W)
        
        # 입력 영역
        input_frame = ttk.Frame(main_frame)
        input_frame.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(input_frame, text="행 선택:").pack(anchor=tk.W)
        
        entry_var = tk.StringVar()
        entry = ttk.Entry(input_frame, textvariable=entry_var, font=('맑은 고딕', 11))
        entry.pack(fill=tk.X, pady=(5, 0))
        entry.focus()
        
        # 결과 표시 영역
        result_frame = ttk.LabelFrame(main_frame, text="선택 결과", padding="10")
        result_frame.pack(fill=tk.X, pady=(0, 20))
        
        text_frame = ttk.Frame(result_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        result_text = tk.Text(text_frame, height=8, width=50, wrap=tk.WORD, font=('맑은 고딕', 9))
        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=result_text.yview)
        result_text.configure(yscrollcommand=scrollbar.set)
        
        result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 버튼 영역
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        # 이벤트 핸들러 설정
        self._setup_event_handlers(root, entry_var, result_text)
        
        ttk.Button(button_frame, text="확인", 
                  command=lambda: self._confirm_selection(entry_var.get(), root)).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="취소", 
                  command=lambda: self._cancel_selection(root)).pack(side=tk.LEFT)
    
    def _setup_event_handlers(self, root, entry_var, result_text):
        """이벤트 핸들러 설정"""
        def preview_selection():
            selection = entry_var.get()
            if not selection.strip():
                result_text.delete(1.0, tk.END)
                result_text.insert(1.0, "행을 입력하세요.")
                return
            
            try:
                rows = self.parse_row_selection(selection, silent=True)
                if rows:
                    result_text.delete(1.0, tk.END)
                    result_text.insert(1.0, f"선택된 행: {rows}\n")
                    result_text.insert(tk.END, f"총 {len(rows)}개 행이 선택됩니다.\n\n")
                    
                    # 선택된 행의 미리보기
                    self._show_row_preview(result_text, rows)
                else:
                    result_text.delete(1.0, tk.END)
                    result_text.insert(1.0, "올바른 행 번호를 입력하세요.")
            except Exception as e:
                result_text.delete(1.0, tk.END)
                result_text.insert(1.0, f"오류: {e}")
        
        # 실시간 미리보기
        entry_var.trace('w', lambda *args: preview_selection())
        
        # Enter 키로 확인
        root.bind('<Return>', lambda event: self._confirm_selection(entry_var.get(), root))
    
    def _show_row_preview(self, result_text, rows):
        """선택된 행의 미리보기 표시"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(self.file_path)
            if self.config.sheet_name in wb.sheetnames:
                ws = wb[self.config.sheet_name]
            else:
                ws = wb.active
            
            max_row = ws.max_row
            result_text.insert(tk.END, f"시트 최대 행: {max_row}\n")
            
            # 각 행의 상호명 표시 (4번째 컬럼 기준)
            for row_num in rows[:5]:  # 최대 5개만 미리보기
                if row_num <= max_row:
                    company_value = ws.cell(row=row_num, column=4).value or "데이터 없음"
                    result_text.insert(tk.END, f"행{row_num}: {company_value}\n")
                else:
                    result_text.insert(tk.END, f"행{row_num}: 범위 초과\n")
            
            if len(rows) > 5:
                result_text.insert(tk.END, f"... 외 {len(rows)-5}개 행\n")
                
        except Exception as e:
            result_text.insert(tk.END, f"미리보기 실패: {e}")
    
    def parse_row_selection(self, selection: str, silent: bool = False) -> List[int]:
        """행 선택 문자열 파싱"""
        if not selection.strip():
            return []
        
        rows = []
        parts = selection.split(',')
        
        for part in parts:
            part = part.strip()
            if '-' in part:
                # 범위 처리 (예: 2-8)
                try:
                    start_str, end_str = part.split('-', 1)
                    if start_str and end_str:
                        start_num = int(start_str.strip())
                        end_num = int(end_str.strip())
                        rows.extend(range(start_num, end_num + 1))
                    else:
                        if not silent:
                            print(f"❌ 잘못된 범위 형식: {part}")
                except ValueError:
                    if not silent:
                        print(f"❌ 잘못된 범위 형식: {part}")
            else:
                # 단일 행 처리
                try:
                    row_num = int(part.strip())
                    rows.append(row_num)
                except ValueError:
                    if not silent:
                        print(f"❌ 잘못된 행 번호: {part}")
        
        return sorted(set(rows))  # 중복 제거 및 정렬
    
    def _confirm_selection(self, selection: str, root):
        """선택 확정"""
        rows = self.parse_row_selection(selection)
        
        if not rows:
            messagebox.showerror("오류", "올바른 행을 선택하세요.")
            return
        
        self.selected_rows = rows
        root.destroy()
    
    def _cancel_selection(self, root):
        """선택 취소"""
        self.selected_rows = None
        root.destroy()
        print("사용자가 프로그램을 취소했습니다.")
        sys.exit(0)


class DataProcessor:
    """데이터 처리 클래스"""
    
    def __init__(self, config: SheetConfig, file_path: str):
        self.config = config
        self.file_path = file_path
        self.headers = None
        self.processed_data = []
    
    def process_excel_data(self, selected_rows: List[int]) -> bool:
        """엑셀 데이터 처리"""
        if not self.file_path or not selected_rows:
            print("❌ 엑셀 파일 경로나 선택된 행이 없습니다.")
            return False
        
        try:
            # openpyxl로 실제 사용된 범위 확인
            from openpyxl import load_workbook
            wb = load_workbook(self.file_path)
            
            if self.config.sheet_name in wb.sheetnames:
                ws = wb[self.config.sheet_name]
            else:
                ws = wb.active
                print(f"경고: '{self.config.sheet_name}' 시트를 찾을 수 없어 기본 시트({ws.title}) 사용")
            
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"시트 정보: {max_row}행 × {max_col}열")
            
            # pandas로 데이터 읽기
            df = pd.read_excel(self.file_path, sheet_name=self.config.sheet_name, header=None, 
                             dtype=str, keep_default_na=False, engine='openpyxl', na_filter=False, nrows=max_row)
            print(f"데이터 읽기 완료: {len(df)}행 × {len(df.columns)}열")
            
            # 헤더 설정
            if len(df) < 1:
                print("❌ 엑셀 파일에 데이터가 없습니다.")
                return False
            
            self.headers = [str(h).strip() for h in df.iloc[0].fillna("").tolist()]
            print(f"헤더: {self.headers}")
            
            # 선택된 행들 처리
            self.processed_data = []
            for row_num in selected_rows:
                if row_num > len(df):
                    print(f"⚠️ 행 {row_num}은 데이터 범위({len(df)})를 초과합니다.")
                    continue
                
                # 행 데이터 추출
                row_data = df.iloc[row_num - 1].fillna("").tolist()
                
                # 헤더와 데이터 매핑
                row_dict = {}
                for i, header in enumerate(self.headers):
                    if i < len(row_data):
                        value = str(row_data[i]).strip()
                        row_dict[header] = self._process_field_data(header, value)
                    else:
                        row_dict[header] = ""
                
                self.processed_data.append({
                    'row_number': row_num,
                    'data': row_dict
                })
                
                print(f"✅ 행 {row_num} 처리 완료")
            
            print(f"✅ 총 {len(self.processed_data)}개 행 처리 완료")
            return True
            
        except Exception as e:
            print(f"❌ 엑셀 데이터 처리 실패: {e}")
            return False
    
    def _process_field_data(self, header: str, value: str) -> str:
        """필드별 데이터 처리"""
        # 사업자번호 처리 (하이픈 제거)
        if any(keyword in header for keyword in ['사업자번호', '사업자등록번호', '거래처등록번호', '등록번호']):
            return ''.join(filter(str.isdigit, value))
        
        # 이메일 처리
        elif '이메일' in header:
            if '@' in value:
                parts = value.split('@', 1)
                return {'front': parts[0].strip(), 'back': parts[1].strip()}
            else:
                return {'front': value, 'back': ''}
        
        return value
    
    def get_processed_data(self) -> List[Dict]:
        """처리된 데이터 반환"""
        return self.processed_data


class StatusRecorder:
    """상태 기록 클래스"""
    
    def __init__(self, config: SheetConfig, file_path: str):
        self.config = config
        self.file_path = file_path
    
    def write_success(self, row_number: int, message: str = None) -> bool:
        """성공 상태 기록"""
        if message is None:
            message = datetime.now().strftime("%Y-%m-%d")
        
        return self._write_to_excel(row_number, message)
    
    def write_error(self, row_number: int, error_message: str = "error") -> bool:
        """에러 상태 기록"""
        return self._write_to_excel(row_number, error_message)
    
    def write_error_to_matching_business_numbers(self, business_number: str, error_message: str = "번호오류") -> bool:
        """같은 사업자번호의 모든 행에 에러 기록"""
        try:
            df = pd.read_excel(self.file_path, sheet_name=self.config.sheet_name)
            
            # 같은 등록번호를 가진 모든 행 찾기
            target_business_number = str(business_number).replace('-', '').strip()
            matching_rows = []
            
            for idx, row in df.iterrows():
                row_business_number = str(row.get(self.config.business_number_column, '')).replace('-', '').strip()
                if row_business_number == target_business_number:
                    excel_row_number = idx + 2  # pandas index 0부터, 엑셀 1부터, 헤더 고려해서 +2
                    matching_rows.append(excel_row_number)
            
            if not matching_rows:
                print(f"❌ 등록번호 {business_number}와 일치하는 행을 찾을 수 없습니다.")
                return False
            
            print(f"발견된 일치 행들: {matching_rows}")
            
            # 모든 일치 행에 에러 기록
            success_count = 0
            for row_num in matching_rows:
                if self._write_to_excel(row_num, error_message):
                    success_count += 1
            
            print(f"✅ {success_count}/{len(matching_rows)}개 행에 에러 기록 완료")
            return success_count > 0
            
        except Exception as e:
            print(f"❌ 같은 등록번호 에러 기록 실패: {e}")
            return False
    
    def _write_to_excel(self, row_number: int, message: str) -> bool:
        """엑셀에 메시지 기록"""
        # 방법 1: xlwings로 열린 파일에 직접 쓰기
        if self._write_with_xlwings(row_number, message):
            return True
        
        # 방법 2: openpyxl로 파일 수정
        return self._write_with_openpyxl(row_number, message)
    
    def _write_with_xlwings(self, row_number: int, message: str) -> bool:
        """xlwings로 열린 엑셀 파일에 기록"""
        try:
            import xlwings as xw
            
            app = None
            try:
                app = xw.apps.active if xw.apps else None
            except:
                pass
            
            if not app:
                try:
                    app = xw.App(visible=True, add_book=False)
                except:
                    return False
            
            if not (app and hasattr(app, 'books')):
                return False
            
            # 열린 워크북 찾기
            workbook_name = os.path.basename(self.file_path)
            wb = None
            
            for book in app.books:
                if book.name == workbook_name:
                    wb = book
                    break
            
            if not wb:
                return False
            
            # 해당 시트 선택
            ws = None
            for sheet in wb.sheets:
                if sheet.name == self.config.sheet_name:
                    ws = sheet
                    break
            
            if not ws:
                ws = wb.sheets[0]  # 첫 번째 시트 사용
            
            # 지정된 컬럼에 메시지 기록
            col_letter = self._get_column_letter(self.config.status_column)
            ws.range(f'{col_letter}{row_number}').value = message
            wb.save()
            
            print(f"✅ 행 {row_number} {col_letter}열에 '{message}' 기록 완료 (xlwings)")
            return True
            
        except ImportError:
            print("   xlwings가 설치되지 않았습니다.")
            return False
        except Exception as e:
            print(f"   xlwings 기록 실패: {e}")
            return False
    
    def _write_with_openpyxl(self, row_number: int, message: str) -> bool:
        """openpyxl로 파일에 직접 기록"""
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(self.file_path)
            
            if self.config.sheet_name in workbook.sheetnames:
                worksheet = workbook[self.config.sheet_name]
            else:
                worksheet = workbook.active
            
            # 지정된 컬럼에 메시지 기록
            worksheet.cell(row=row_number, column=self.config.status_column, value=message)
            
            workbook.save(self.file_path)
            workbook.close()
            
            col_letter = self._get_column_letter(self.config.status_column)
            print(f"✅ 행 {row_number} {col_letter}열에 '{message}' 기록 완료 (openpyxl)")
            return True
            
        except Exception as e:
            print(f"❌ openpyxl 기록 실패: {e}")
            return False
    
    def _get_column_letter(self, column_number: int) -> str:
        """컬럼 번호를 문자로 변환 (1=A, 2=B, ..., 17=Q)"""
        if column_number <= 26:
            return chr(64 + column_number)  # A=65
        else:
            # 26을 넘는 경우 (AA, AB 등)
            first = (column_number - 1) // 26
            second = (column_number - 1) % 26 + 1
            return chr(64 + first) + chr(64 + second)


class ExcelUnifiedProcessor:
    """엑셀 데이터 통합 처리 메인 클래스"""
    
    def __init__(self, sheet_type: str = "partner"):
        """
        Args:
            sheet_type: "partner" (거래처) 또는 "transaction" (거래명세표)
        """
        if sheet_type == "partner":
            self.config = SheetConfig.get_partner_config()
        elif sheet_type == "transaction":
            self.config = SheetConfig.get_transaction_config()
        else:
            raise ValueError("sheet_type must be 'partner' or 'transaction'")
        
        self.file_manager = ExcelFileManager(self.config)
        self.row_selector = None
        self.data_processor = None
        self.status_recorder = None
        
        self.selected_rows = None
        self.processed_data = []
    
    def initialize(self) -> bool:
        """초기화 - 파일 열기 및 컴포넌트 생성"""
        # 파일 열기
        if not self.file_manager.check_and_open_excel():
            print("❌ 엑셀 파일 열기에 실패했습니다.")
            return False
        
        excel_file_path = self.file_manager.excel_file_path
        
        # 컴포넌트 초기화
        self.row_selector = RowSelector(self.config, excel_file_path)
        self.data_processor = DataProcessor(self.config, excel_file_path)
        self.status_recorder = StatusRecorder(self.config, excel_file_path)
        
        return True
    
    def select_rows(self) -> bool:
        """행 선택"""
        if not self.row_selector:
            print("❌ row_selector가 초기화되지 않았습니다.")
            return False
        
        if not self.row_selector.show_row_selection_gui():
            print("❌ 행 선택이 취소되었습니다.")
            return False
        
        self.selected_rows = self.row_selector.selected_rows
        print(f"✅ 선택된 행: {self.selected_rows}")
        return True
    
    def process_data(self) -> bool:
        """데이터 처리"""
        if not self.data_processor or not self.selected_rows:
            print("❌ data_processor나 selected_rows가 없습니다.")
            return False
        
        if not self.data_processor.process_excel_data(self.selected_rows):
            print("❌ 데이터 처리에 실패했습니다.")
            return False
        
        self.processed_data = self.data_processor.get_processed_data()
        return True
    
    def record_success(self, row_number: int, message: str = None) -> bool:
        """성공 상태 기록"""
        if not self.status_recorder:
            print("❌ status_recorder가 초기화되지 않았습니다.")
            return False
        
        return self.status_recorder.write_success(row_number, message)
    
    def record_error(self, row_number: int, error_message: str = "error") -> bool:
        """에러 상태 기록"""
        if not self.status_recorder:
            print("❌ status_recorder가 초기화되지 않았습니다.")
            return False
        
        return self.status_recorder.write_error(row_number, error_message)
    
    def record_error_for_business_number(self, business_number: str, error_message: str = "번호오류") -> bool:
        """같은 사업자번호의 모든 행에 에러 기록"""
        if not self.status_recorder:
            print("❌ status_recorder가 초기화되지 않았습니다.")
            return False
        
        return self.status_recorder.write_error_to_matching_business_numbers(business_number, error_message)
    
    def get_processed_data(self) -> List[Dict]:
        """처리된 데이터 반환"""
        return self.processed_data
    
    def get_selected_rows(self) -> List[int]:
        """선택된 행 반환"""
        return self.selected_rows or []


# 편의 함수
def create_partner_processor() -> ExcelUnifiedProcessor:
    """거래처 시트용 프로세서 생성"""
    return ExcelUnifiedProcessor("partner")

def create_transaction_processor() -> ExcelUnifiedProcessor:
    """거래명세표 시트용 프로세서 생성"""
    return ExcelUnifiedProcessor("transaction")


if __name__ == "__main__":
    # 테스트 코드
    print("=== 엑셀 통합 처리 모듈 테스트 ===")
    
    # 거래처 시트 테스트
    processor = create_partner_processor()
    
    if processor.initialize():
        print("✅ 초기화 완료")
        
        if processor.select_rows():
            print("✅ 행 선택 완료")
            
            if processor.process_data():
                print("✅ 데이터 처리 완료")
                
                # 처리된 데이터 확인
                data = processor.get_processed_data()
                print(f"처리된 데이터: {len(data)}개")
                
                if data:
                    print("첫 번째 행 데이터 샘플:")
                    for key, value in list(data[0]['data'].items())[:3]:
                        print(f"  {key}: {value}")
            else:
                print("❌ 데이터 처리 실패")
        else:
            print("❌ 행 선택 실패")
    else:
        print("❌ 초기화 실패")