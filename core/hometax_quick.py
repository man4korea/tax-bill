#-*- coding: utf-8 -*-
import asyncio
import os
import sys
import subprocess
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime
from dotenv import load_dotenv
from playwright.async_api import async_playwright
from excel_data_manager import ExcelDataManager
import pandas as pd
import winsound  # Windows Beep 처리용

# 거래 내역 입력 프로세스 모듈 import
from hometax_transaction_processor import (
    process_transaction_details,
    get_same_business_number_rows,
    check_and_update_supply_date,
    input_transaction_items_basic,
    input_transaction_items_extended,
    input_single_transaction_item,
    finalize_transaction_summary,
    verify_and_calculate_credit,
    show_amount_mismatch_dialog,
    handle_issuance_alerts,
    write_to_tax_invoice_sheet,
    clear_form_fields
)

async def play_beep(count=1, frequency=800, duration=300):
    """지정된 횟수만큼 Beep음을 재생합니다."""
    try:
        print(f"      🔊 Beep 알림 {count}회...")
        for i in range(count):
            winsound.Beep(frequency, duration)
            if i < count - 1:
                await asyncio.sleep(0.2)
        print("      🔊 Beep 알림 완료")
    except Exception as beep_error:
        print(f"      Beep 처리 오류: {beep_error}")

class TaxInvoiceExcelProcessor:
    def __init__(self):
        self.selected_rows = None
        self.selected_data = None
        self.excel_file_path = None
        self.headers = None
        
        # 엑셀 거래명세표 컬럼과 홈택스 필드 매칭 테이블
        self.field_mapping = {
            # 엑셀 컬럼 인덱스: (엑셀 컬럼명, 홈택스 필드 ID, 데이터 처리 함수)
            0: ('작성일자', 'supply_date', self._format_date),
            1: ('등록번호', 'business_number', self._format_business_number), 
            2: ('상호', 'company_name', str),
            3: ('품목코드', 'item_code', str),
            4: ('품명', 'item_name', str),
            5: ('규격', 'spec', str),
            6: ('수량', 'quantity', self._format_number),
            7: ('단가', 'unit_price', self._format_number),
            8: ('공급가액', 'supply_amount', self._format_number),
            9: ('세액', 'tax_amount', self._format_number)
        }
        
        # HomeTax 실제 필드 선택자 (기본 정보)
        self.base_selectors = {
            'business_number': '#mf_txppWframe_edtDmnrBsnoTop',
            'business_number_confirm': '#mf_txppWframe_btnDmnrBsnoCnfrTop',
            'company_name': '#mf_txppWframe_edtDmnrTnmNmTop',
            'representative_name': '#mf_txppWframe_edtDmnrRprsFnmTop',
            'email_id': '#mf_txppWframe_edtDmnrMchrgEmlIdTop',
            'email_domain': '#mf_txppWframe_edtDmnrMchrgEmlDmanTop',
            'supply_date': '#mf_txppWframe_calWrtDtTop_input',
            'add_item': '#mf_txppWframe_btnLsatAddTop',
            'delete_item': '#mf_txppWframe_btnLsatDltTop',
            'total_amount': '#mf_txppWframe_edtTotaAmtHeaderTop',
            'total_supply': '#mf_txppWframe_edtSumSplCftHeaderTop',
            'total_tax': '#mf_txppWframe_edtSumTxamtHeaderTop',
            'issue_button': '#mf_txppWframe_btnIsn',
            'hold_button': '#mf_txppWframe_btnIsnRsrv'
        }
        
        # HomeTax 품목별 필드 선택자 템플릿 (row_idx로 동적 생성)
        self.item_selectors = {
            'supply_date': '#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatSplDdTop',
            'item_name': '#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatNmTop',
            'spec': '#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatRszeNmTop',
            'quantity': '#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatQtyTop',
            'unit_price': '#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatUtprcTop',
            'supply_amount': '#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatSplCftTop',
            'tax_amount': '#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatTxamtTop'
        }
    
    def _format_date(self, value):
        """날짜 형식 변환 (YYYY-MM-DD → YYYYMMDD)"""
        if pd.isna(value) or not value:
            return ""
        date_str = str(value).replace('-', '').replace('/', '').replace('.', '')
        return date_str[:8] if len(date_str) >= 8 else date_str
    
    def _format_business_number(self, value):
        """사업자번호 형식 변환 (하이픈 제거)"""
        if pd.isna(value) or not value:
            return ""
        return str(value).replace('-', '').strip()
    
    def _format_number(self, value):
        """숫자 형식 변환"""
        if pd.isna(value) or not value:
            return "0"
        try:
            if isinstance(value, str):
                # 콤마 제거 후 숫자로 변환
                clean_value = value.replace(',', '').strip()
                return str(int(float(clean_value)))
            else:
                return str(int(float(value)))
        except:
            return "0"
    
    def write_error_to_excel(self, row_number, error_message="error"):
        """엑셀 파일의 지정된 행 발행일 열에 에러 메시지 작성"""
        if not self.excel_file_path:
            print("❌ 엑셀 파일 경로가 없습니다.")
            return False
        
        try:
            from openpyxl import load_workbook
            
            print(f"엑셀 파일에 에러 기록 중: 행 {row_number}, 메시지: {error_message}")
            
            workbook = load_workbook(self.excel_file_path)
            
            # 거래명세표 시트 선택
            if "거래명세표" in workbook.sheetnames:
                worksheet = workbook["거래명세표"]
            else:
                worksheet = workbook.active
                print(f"경고: '거래명세표' 시트를 찾을 수 없어 기본 시트({worksheet.title}) 사용")
            
            # 발행일 열(첫 번째 열)에 에러 메시지 작성
            worksheet.cell(row=row_number, column=1, value=error_message)
            
            workbook.save(self.excel_file_path)
            workbook.close()
            
            print(f"✅ 엑셀 파일에 에러 기록 완료: 행 {row_number}")
            return True
            
        except Exception as e:
            print(f"❌ 엑셀 파일 에러 기록 실패: {e}")
            return False
    
    def write_error_to_excel_q_column(self, row_number, error_message="번호오류"):
        """엑셀 파일의 거래명세표 시트 Q열(발행일)에 에러 메시지 작성 (단일 행)"""
        if not self.excel_file_path:
            print("❌ 엑셀 파일 경로가 없습니다.")
            return False
        
        try:
            from openpyxl import load_workbook
            
            print(f"엑셀 Q열에 에러 기록 중: 행 {row_number}, 메시지: {error_message}")
            
            workbook = load_workbook(self.excel_file_path)
            
            # 거래명세표 시트 선택
            if "거래명세표" in workbook.sheetnames:
                worksheet = workbook["거래명세표"]
            else:
                worksheet = workbook.active
                print(f"경고: '거래명세표' 시트를 찾을 수 없어 기본 시트({worksheet.title}) 사용")
            
            # Q열(17번째 열)에 에러 메시지 작성 (Q = 17번째 컬럼)
            worksheet.cell(row=row_number, column=17, value=error_message)
            
            workbook.save(self.excel_file_path)
            workbook.close()
            
            print(f"✅ 엑셀 Q열에 에러 기록 완료: 행 {row_number}, Q열: {error_message}")
            return True
            
        except Exception as e:
            print(f"❌ 엑셀 Q열 에러 기록 실패: {e}")
            return False
    
    def write_completion_to_excel_q_column(self, row_number, completion_message="완료"):
        """엑셀 파일의 거래명세표 시트 Q열(발행일)에 완료 메시지 작성 (단일 행)"""
        if not self.excel_file_path:
            print("❌ 엑셀 파일 경로가 없습니다.")
            return False
        
        try:
            from openpyxl import load_workbook
            import time
            
            print(f"엑셀 Q열에 완료 기록 중: 행 {row_number}, 메시지: {completion_message}")
            
            # 파일이 열려있는 경우를 대비해 여러 번 시도
            max_attempts = 3
            for attempt in range(max_attempts):
                try:
                    workbook = load_workbook(self.excel_file_path)
                    
                    # 거래명세표 시트 선택
                    if "거래명세표" in workbook.sheetnames:
                        worksheet = workbook["거래명세표"]
                    else:
                        worksheet = workbook.active
                        print(f"경고: '거래명세표' 시트를 찾을 수 없어 기본 시트({worksheet.title}) 사용")
                    
                    # Q열 (17번째 컬럼)에 완료 메시지 작성
                    worksheet.cell(row=row_number, column=17, value=completion_message)
                    
                    workbook.save(self.excel_file_path)
                    workbook.close()
                    
                    print(f"   ✅ 행 {row_number} Q열에 '{completion_message}' 완료 기록 (openpyxl)")
                    return True
                    
                except PermissionError as pe:
                    if attempt < max_attempts - 1:
                        print(f"   ⚠️ Excel 파일이 사용 중입니다. {attempt + 1}/{max_attempts} 시도 후 재시도...")
                        time.sleep(1)  # 1초 대기 후 재시도
                        continue
                    else:
                        print(f"   ❌ Excel 파일 권한 오류 (파일이 열려있음): {pe}")
                        return False
                except Exception as inner_e:
                    if attempt < max_attempts - 1:
                        print(f"   ⚠️ Excel 작업 오류, {attempt + 1}/{max_attempts} 재시도 중: {inner_e}")
                        time.sleep(0.5)
                        continue
                    else:
                        raise inner_e
            
            return False
            
        except Exception as e:
            print(f"❌ 엑셀 Q열 완료 기록 실패: {e}")
            print("   💡 Excel 파일이 다른 프로그램에서 열려있지 않은지 확인해주세요.")
            return False
    
    def write_error_to_all_matching_business_numbers(self, business_number, error_message="번호오류"):
        """같은 사업자등록번호를 가진 모든 행의 Q열에 에러 메시지 작성"""
        if not self.excel_file_path:
            print("❌ 엑셀 파일 경로가 없습니다.")
            return False
        
        try:
            import pandas as pd
            
            print(f"같은 등록번호({business_number})를 가진 모든 행에 Q열 에러 기록 중...")
            
            # pandas로 데이터 읽기 (행 찾기용)
            try:
                df = pd.read_excel(self.excel_file_path, sheet_name='거래명세표')
            except:
                df = pd.read_excel(self.excel_file_path)  # 기본 시트 사용
            
            # 사업자등록번호 형식 통일 (하이픈 제거)
            target_business_number = str(business_number).replace('-', '').strip()
            
            # 같은 등록번호를 가진 모든 행 찾기
            matching_rows = []
            for idx, row in df.iterrows():
                row_business_number = str(row.get('등록번호', '')).replace('-', '').strip()
                if row_business_number == target_business_number:
                    excel_row_number = idx + 2  # pandas index는 0부터, 엑셀은 1부터, 헤더 고려하면 +2
                    matching_rows.append(excel_row_number)
            
            if not matching_rows:
                print(f"❌ 등록번호 {business_number}와 일치하는 행을 찾을 수 없습니다.")
                return False
            
            print(f"발견된 일치 행들: {matching_rows} (총 {len(matching_rows)}개)")
            
            # 방법 1: xlwings를 사용해서 열린 엑셀 파일에 직접 쓰기 시도
            try:
                import xlwings as xw
                
                # 현재 열려있는 엑셀 앱에 연결
                try:
                    app = xw.apps.active
                except:
                    app = xw.App(visible=True, add_book=False)
                
                # 열린 워크북 찾기
                workbook_name = self.excel_file_path.split("\\")[-1]  # 파일명만 추출
                wb = None
                
                for book in app.books:
                    if book.name == workbook_name:
                        wb = book
                        break
                
                if wb:
                    # "거래명세표" 시트 선택
                    ws = None
                    for sheet in wb.sheets:
                        if sheet.name == "거래명세표":
                            ws = sheet
                            break
                    
                    if not ws:
                        ws = wb.sheets[0]  # 첫 번째 시트 사용
                    
                    # Q열(17번째 열)에 에러 메시지 기록
                    updated_count = 0
                    for row_number in matching_rows:
                        try:
                            ws.range(f'Q{row_number}').value = error_message
                            updated_count += 1
                            print(f"   행 {row_number} Q열에 '{error_message}' 작성 완료 (xlwings)")
                        except Exception as e:
                            print(f"   행 {row_number} Q열 작성 실패 (xlwings): {e}")
                            continue
                    
                    # 저장
                    wb.save()
                    
                    print(f"✅ 등록번호 {business_number}의 모든 행 Q열 에러 기록 완료 (xlwings): {updated_count}/{len(matching_rows)}개 행")
                    return True
                    
            except ImportError:
                print("   xlwings가 설치되지 않았습니다. openpyxl 방법을 시도합니다...")
            except Exception as e:
                print(f"   xlwings 방법 실패: {e}")
            
            # 방법 2: openpyxl로 파일 직접 수정 (엑셀이 닫혀있을 때만 가능)
            from openpyxl import load_workbook
            
            workbook = load_workbook(self.excel_file_path)
            
            # 거래명세표 시트 선택
            if "거래명세표" in workbook.sheetnames:
                worksheet = workbook["거래명세표"]
            else:
                worksheet = workbook.active
                print(f"경고: '거래명세표' 시트를 찾을 수 없어 기본 시트({worksheet.title}) 사용")
            
            # 모든 일치하는 행의 Q열(17번째 열)에 에러 메시지 작성
            updated_count = 0
            for row_number in matching_rows:
                try:
                    worksheet.cell(row=row_number, column=17, value=error_message)
                    updated_count += 1
                    print(f"   행 {row_number} Q열에 '{error_message}' 작성 완료 (openpyxl)")
                except Exception as e:
                    print(f"   행 {row_number} Q열 작성 실패 (openpyxl): {e}")
                    continue
            
            # 파일 저장
            workbook.save(self.excel_file_path)
            workbook.close()
            
            print(f"✅ 등록번호 {business_number}의 모든 행 Q열 에러 기록 완료 (openpyxl): {updated_count}/{len(matching_rows)}개 행")
            return True
            
        except PermissionError as pe:
            print(f"❌ 파일 권한 오류: {pe}")
            print("   🔧 해결 방법:")
            print("   1. 엑셀 파일이 열려있다면 파일을 닫고 다시 시도하세요")
            print("   2. 또는 xlwings를 설치하세요: pip install xlwings")
            return False
            
        except Exception as e:
            print(f"❌ 같은 등록번호 모든 행 Q열 에러 기록 실패: {e}")
            return False
    
    def write_tax_invoice_data(self, tax_invoice_data):
        """세금계산서 시트에 데이터 기록"""
        if not self.excel_file_path:
            print("❌ 엑셀 파일 경로가 없습니다.")
            return False
        
        try:
            from openpyxl import load_workbook
            
            print(f"세금계산서 시트에 데이터 기록 중...")
            
            # 방법 1: xlwings를 사용해서 열린 엑셀 파일에 직접 쓰기 시도
            try:
                import xlwings as xw
                
                # 현재 열려있는 엑셀 앱에 연결
                try:
                    app = xw.apps.active
                except:
                    app = xw.App(visible=True, add_book=False)
                
                # 열린 워크북 찾기
                workbook_name = self.excel_file_path.split("\\")[-1]
                wb = None
                
                for book in app.books:
                    if book.name == workbook_name:
                        wb = book
                        break
                
                if wb:
                    # "세금계산서" 시트 찾기 또는 생성
                    ws = None
                    for sheet in wb.sheets:
                        if sheet.name == "세금계산서":
                            ws = sheet
                            break
                    
                    if not ws:
                        # 세금계산서 시트가 없으면 생성
                        ws = wb.sheets.add("세금계산서")
                        # 헤더 작성
                        headers = ['공급일자', '등록번호', '상호', '이메일', '', '품목', '규격', '수량', '공급가액', '세액', '합계금액', '기간및건수']
                        for i, header in enumerate(headers, 1):
                            ws.range(f'{chr(64+i)}1').value = header
                    
                    # 마지막 행 찾기
                    last_row = 1
                    while ws.range(f'A{last_row}').value is not None:
                        last_row += 1
                    
                    # 데이터 기록
                    for col_letter, value in tax_invoice_data.items():
                        if value:  # 값이 있을 때만 기록
                            ws.range(f'{col_letter.upper()}{last_row}').value = value
                    
                    # 저장
                    wb.save()
                    
                    print(f"✅ 세금계산서 시트에 데이터 기록 완료 (xlwings): 행 {last_row}")
                    return True
                    
            except ImportError:
                print("   xlwings가 설치되지 않았습니다. openpyxl 방법을 시도합니다...")
            except Exception as e:
                print(f"   xlwings 방법 실패: {e}")
            
            # 방법 2: openpyxl로 파일 직접 수정
            workbook = load_workbook(self.excel_file_path)
            
            # 세금계산서 시트 찾기 또는 생성
            if "세금계산서" in workbook.sheetnames:
                worksheet = workbook["세금계산서"]
            else:
                worksheet = workbook.create_sheet("세금계산서")
                # 헤더 작성
                headers = ['공급일자', '등록번호', '상호', '이메일', '', '품목', '규격', '수량', '공급가액', '세액', '합계금액', '기간및건수']
                for i, header in enumerate(headers, 1):
                    worksheet.cell(row=1, column=i, value=header)
            
            # 마지막 행 찾기
            last_row = 1
            while worksheet.cell(row=last_row, column=1).value is not None:
                last_row += 1
            
            # 컬럼 매핑 (a=1, b=2, c=3, ...)
            column_mapping = {
                'a': 1, 'b': 2, 'c': 3, 'd': 4, 'e': 5, 'f': 6, 'g': 7, 'h': 8,
                'i': 9, 'j': 10, 'k': 11, 'l': 12
            }
            
            # 데이터 기록
            for col_letter, value in tax_invoice_data.items():
                if value and col_letter.lower() in column_mapping:
                    col_num = column_mapping[col_letter.lower()]
                    worksheet.cell(row=last_row, column=col_num, value=value)
            
            # 파일 저장
            workbook.save(self.excel_file_path)
            workbook.close()
            
            print(f"✅ 세금계산서 시트에 데이터 기록 완료 (openpyxl): 행 {last_row}")
            return True
            
        except PermissionError as pe:
            print(f"❌ 파일 권한 오류: {pe}")
            print("   🔧 해결 방법:")
            print("   1. 엑셀 파일이 열려있다면 파일을 닫고 다시 시도하세요")
            print("   2. 또는 xlwings를 설치하세요: pip install xlwings")
            return False
            
        except Exception as e:
            print(f"❌ 세금계산서 시트 기록 실패: {e}")
            return False
    
    def check_and_open_excel_file(self):
        """세금계산서.xlsx 파일 체크 및 자동 열기 (hometax_excel_integration.py와 동일)"""
        target_filename = "세금계산서.xlsx"
        
        # OneDrive 문서 폴더와 일반 문서 폴더 둘 다 확인
        onedrive_documents = os.path.expanduser("~/OneDrive/문서")
        regular_documents = os.path.expanduser("~/Documents")
        
        # 먼저 OneDrive 문서 폴더 확인
        if os.path.exists(os.path.join(onedrive_documents, target_filename)):
            target_file = os.path.join(onedrive_documents, target_filename)
            documents_path = onedrive_documents
        elif os.path.exists(os.path.join(regular_documents, target_filename)):
            target_file = os.path.join(regular_documents, target_filename)
            documents_path = regular_documents
        else:
            # 둘 다 없으면 OneDrive 폴더를 기본으로 사용
            target_file = os.path.join(onedrive_documents, target_filename)
            documents_path = onedrive_documents
        
        print(f"\n=== 엑셀 파일 체크 및 열기 ===")
        
        # === 1단계: 세금계산서.xlsx가 이미 열려있는가? ===
        print(f"1단계: '{target_filename}'가 이미 열려있는지 확인...")
        try:
            import psutil
            excel_processes = [p for p in psutil.process_iter(['pid', 'name']) if 'excel' in p.info['name'].lower()]
            if excel_processes:
                print("   Excel 프로세스 실행 중")
                
                # xlwings로 정확한 파일 확인
                try:
                    import xlwings as xw
                    app = xw.App(visible=True, add_book=False)
                    
                    if app.books:
                        print(f"   열린 Excel 파일들을 확인합니다...")
                        for book in app.books:
                            print(f"   - 확인 중: '{book.name}'")
                            if book.name.lower() == target_filename.lower():
                                print(f"   ✅ '{book.name}' 파일이 이미 열려있습니다! 중복 열기 방지")
                                self.excel_file_path = book.fullname
                                return True
                            elif target_filename.lower().replace('.xlsx', '') in book.name.lower():
                                print(f"   유사한 파일명 발견: '{book.name}' (읽기 전용일 수 있음)")
                                # 유사한 파일명도 이미 열려있는 것으로 처리
                                self.excel_file_path = book.fullname
                                print(f"   ✅ 유사 파일 사용: '{book.name}' - 중복 열기 방지")
                                return True
                        print(f"   Excel은 실행 중이지만 '{target_filename}' 파일이 열려있지 않습니다.")
                    else:
                        print("   Excel은 실행 중이지만 열린 파일이 없습니다.")
                        
                except ImportError:
                    print("   xlwings가 설치되지 않았습니다.")
                    print("   xlwings 설치하면 자동 감지 가능: pip install xlwings")
                except Exception as e:
                    print(f"   xlwings 확인 중 오류: {e}")
            else:
                print("   Excel 프로세스가 실행되지 않았습니다.")
        except Exception as e:
            print(f"   프로세스 확인 중 오류: {e}")
        
        # === 2단계: 문서 폴더에 세금계산서.xlsx가 있는가? (1단계에서 찾지 못한 경우만) ===
        print(f"2단계: 문서 폴더에 '{target_filename}' 파일이 있는지 확인...")
        
        if os.path.exists(target_file):
            print(f"   파일 발견: {target_file}")
            
            # 엑셀이 실행 중인지 다시 한번 확인 (1단계에서 놓쳤을 수 있음)
            try:
                import psutil
                excel_processes = [p for p in psutil.process_iter(['pid', 'name']) if 'excel' in p.info['name'].lower()]
                if excel_processes:
                    print(f"   ⚠️ Excel이 실행 중입니다. 중복 열기를 방지하기 위해 파일 경로만 저장합니다.")
                    self.excel_file_path = target_file
                    return True
            except:
                pass
            
            print(f"   '{target_filename}' 파일을 자동으로 엽니다...")
            
            try:
                os.startfile(target_file)
                self.excel_file_path = target_file
                
                # Excel 로딩 대기
                import time
                time.sleep(3)
                
                # 포커스 복원
                try:
                    import win32gui
                    console_hwnd = win32gui.GetConsoleWindow()
                    if console_hwnd:
                        win32gui.SetForegroundWindow(console_hwnd)
                        print("   포커스를 콘솔로 복원")
                except:
                    pass
                
                print(f"   '{target_filename}' 파일이 열렸습니다!")
                return True
                
            except Exception as e:
                print(f"   파일 열기 실패: {e}")
        else:
            print(f"   문서 폴더에 '{target_filename}' 파일이 없습니다.")
        
        # === 3단계: 파일 열기 창으로 세금계산서.xlsx 선택 ===
        print(f"3단계: 파일 선택 창에서 '{target_filename}' 파일을 선택해주세요...")
        
        root = tk.Tk()
        root.withdraw()
        
        file_path = filedialog.askopenfilename(
            title=f"'{target_filename}' 파일을 선택하세요",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=documents_path
        )
        
        if file_path:
            print(f"   선택된 파일: {file_path}")
            
            # 선택한 파일도 Excel이 실행 중이면 중복 열기 방지
            try:
                import psutil
                excel_processes = [p for p in psutil.process_iter(['pid', 'name']) if 'excel' in p.info['name'].lower()]
                if excel_processes:
                    print(f"   ⚠️ Excel이 실행 중입니다. 중복 열기를 방지하기 위해 파일 경로만 저장합니다.")
                    self.excel_file_path = file_path
                    root.destroy()
                    return True
            except:
                pass
            
            try:
                os.startfile(file_path)
                self.excel_file_path = file_path
                
                # 포커스 복원
                import time
                time.sleep(3)
                try:
                    import win32gui
                    console_hwnd = win32gui.GetConsoleWindow()
                    if console_hwnd:
                        win32gui.SetForegroundWindow(console_hwnd)
                        print("   포커스를 콘솔로 복원")
                except:
                    pass
                
                print(f"   파일이 열렸습니다!")
                root.destroy()
                return True
                
            except Exception as e:
                print(f"   파일 열기 실패: {e}")
                root.destroy()
                return False
        else:
            print("   파일이 선택되지 않았습니다.")
            root.destroy()
            return False

    def select_excel_file_and_process(self):
        """엑셀 파일 체크/열기 및 거래명세표 시트에서 행 선택 처리"""
        # 파일 체크 및 자동 열기
        if not self.check_and_open_excel_file():
            print("엑셀 파일 열기에 실패했습니다.")
            return False
        
        # 거래명세표 시트에서 행 선택
        return self.show_row_selection_gui()
    
    def parse_row_selection(self, selection, silent=False):
        """행 선택 문자열을 파싱하여 행 번호 리스트 반환 (hometax_excel_integration.py 방식)"""
        if not selection.strip():
            return []
        
        rows = []
        parts = selection.replace(" ", "").split(",")
        
        for part in parts:
            if "-" in part:
                # 범위 처리 (예: 2-5)
                try:
                    start, end = part.split("-", 1)
                    start_row = int(start)
                    end_row = int(end)
                    if start_row <= end_row:
                        rows.extend(range(start_row, end_row + 1))
                    else:
                        if not silent:
                            print(f"❌ 잘못된 범위: {part}")
                except ValueError:
                    if not silent:
                        print(f"❌ 잘못된 범위: {part}")
            else:
                # 단일 행 처리
                try:
                    row = int(part)
                    if row > 1:  # 헤더 행 제외
                        rows.append(row)
                    else:
                        if not silent:
                            print(f"❌ 잘못된 행 번호: {part}")
                except ValueError:
                    if not silent:
                        print(f"❌ 잘못된 행 번호: {part}")
        
        return sorted(set(rows))  # 중복 제거 및 정렬
    
    def show_row_selection_gui(self):
        """행 선택 GUI 표시 (hometax_excel_integration.py 방식)"""
        print("\n=== 행 선택 GUI ===")
        
        root = tk.Tk()
        root.title("거래명세표 행 선택")
        root.resizable(False, False)
        
        # 화면 중앙에 위치
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        window_width = 600
        window_height = 650
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 4
        root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # 메인 프레임
        main_frame = ttk.Frame(root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 제목
        title_label = ttk.Label(main_frame, text="처리할 거래명세표 행을 선택하세요", 
                               font=('맑은 고딕', 14, 'bold'))
        title_label.pack(pady=(0, 20))
        
        # 안내 메시지
        guide_frame = ttk.LabelFrame(main_frame, text="행 선택 방법 (월 합계 세금계산서)", padding="10")
        guide_frame.pack(fill=tk.X, pady=(0, 20))
        
        guide_text = """
• 단일 행: 2
• 복수 행: 2,4,8
• 범위: 2-8
• 혼합: 2,5-7,10

같은 달의 모든 거래내역을 선택하세요.
거래처별로 16건씩 자동 그룹핑됩니다.

예시: 2행, 5~7행, 10행을 처리하려면 → 2,5-7,10"""
        
        guide_label = ttk.Label(guide_frame, text=guide_text, justify=tk.LEFT)
        guide_label.pack(anchor=tk.W)
        
        # 입력 프레임
        input_frame = ttk.Frame(main_frame)
        input_frame.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(input_frame, text="행 선택:").pack(anchor=tk.W)
        
        entry_var = tk.StringVar()
        entry = ttk.Entry(input_frame, textvariable=entry_var, font=('맑은 고딕', 11))
        entry.pack(fill=tk.X, pady=(5, 0))
        entry.focus()
        
        # 엔터키 이벤트 바인딩
        def on_enter_key(event):
            """엔터키 입력 시 확인 버튼 실행"""
            confirm_selection()
        
        entry.bind('<Return>', on_enter_key)
        
        # 선택 상태 표시 레이블
        status_label = ttk.Label(main_frame, text="행을 입력하고 확인 버튼을 클릭하거나 Enter 키를 누르세요.", 
                                font=('맑은 고딕', 10))
        status_label.pack(pady=(0, 20))
        
        # 버튼 프레임
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        def validate_selection():
            """선택 검증만 수행"""
            selection = entry_var.get()
            if not selection.strip():
                status_label.config(text="행을 입력해주세요.")
                return False
            
            try:
                rows = self.parse_row_selection(selection, silent=True)
                if rows:
                    status_label.config(text=f"총 {len(rows)}개 행이 선택되었습니다.")
                    return True
                else:
                    status_label.config(text="올바른 행 번호를 입력하세요.")
                    return False
            except Exception as e:
                status_label.config(text=f"오류: {e}")
                return False
        
        def confirm_selection():
            """선택 확정"""
            selection = entry_var.get()
            rows = self.parse_row_selection(selection)
            
            if not rows:
                messagebox.showerror("오류", "올바른 행을 선택하세요.")
                return
            
            # 선택된 행 저장
            self.selected_rows = rows
            print(f"선택된 행: {len(rows)}개")
            root.quit()
            root.destroy()
        
        def cancel_selection():
            """선택 취소"""
            self.selected_rows = None
            root.quit()
            root.destroy()
        
        # 실시간 검증
        entry_var.trace('w', lambda *args: validate_selection())
        
        # 버튼 생성 (미리보기 버튼 제거)
        ttk.Button(button_frame, text="확인 (로그인 진행)", command=confirm_selection).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="취소", command=cancel_selection).pack(side=tk.LEFT, padx=5)
        
        root.mainloop()
        
        if self.selected_rows:
            # 거래명세표 시트에서 데이터 추출
            try:
                df = pd.read_excel(self.excel_file_path, sheet_name='거래명세표')
            except:
                df = pd.read_excel(self.excel_file_path)  # 기본 시트 사용
            
            self.selected_data = []
            for row_num in self.selected_rows:
                try:
                    row_data = df.iloc[row_num-2].to_dict()  # -2는 엑셀 행 번호를 pandas 인덱스로 변환
                    row_data['excel_row'] = row_num
                    self.selected_data.append(row_data)
                except IndexError:
                    print(f"경고: 행 {row_num}은 데이터 범위를 벗어났습니다.")
                    continue
            
            print(f"✅ {len(self.selected_data)}개 행의 데이터 추출 완료")
            return True
        
        return False
    
    def group_data_by_business_number(self):
        """사업자번호별로 월 합계 세금계산서 그룹핑 (16건씩)"""
        if not self.selected_data:
            return []
        
        # 1단계: 사업자번호별로만 정렬 (날짜 정렬 제거)
        sorted_data = sorted(self.selected_data, key=lambda x: 
            str(x.get('등록번호', '')).strip()
        )
        
        print(f"✅ 사업자번호별 정렬 완료: {len(sorted_data)}개 행")
        
        # 2단계: 사업자번호별로 월 합계 그룹핑 (16건씩 분할)
        groups = []
        current_business_number = None
        current_group = []
        
        for data in sorted_data:
            business_number = str(data.get('등록번호', '')).strip()
            
            # 새로운 거래처이면 새 그룹 시작
            if business_number != current_business_number:
                if current_group:
                    groups.append(current_group)
                current_group = [data]
                current_business_number = business_number
            # 같은 거래처지만 16건 초과하면 다음 세금계산서로 분할
            elif len(current_group) >= 16:
                groups.append(current_group)
                current_group = [data]
            else:
                current_group.append(data)
        
        # 마지막 그룹 추가
        if current_group:
            groups.append(current_group)
        
        # 그룹 정보 출력 (월 합계 개념)
        print(f"✅ 월 합계 세금계산서 그룹핑 완료: {len(groups)}개 세금계산서")
        
        # 거래처별 세금계산서 개수 요약
        business_summary = {}
        for i, group in enumerate(groups, 1):
            business_number = group[0].get('등록번호', '미상')
            if business_number not in business_summary:
                business_summary[business_number] = []
            business_summary[business_number].append(len(group))
        
        for business_number, invoice_counts in business_summary.items():
            total_items = sum(invoice_counts)
            invoice_count = len(invoice_counts)
            if invoice_count == 1:
                print(f"   거래처 {business_number}: 1장 ({total_items}건)")
            else:
                detail = " + ".join([f"{count}건" for count in invoice_counts])
                print(f"   거래처 {business_number}: {invoice_count}장 ({detail} = 총 {total_items}건)")
        
        return groups
    
    def get_processed_row_data(self, row_index):
        """선택된 행의 데이터를 홈택스 필드용으로 가공하여 반환"""
        if not self.selected_data or row_index >= len(self.selected_data):
            return None
        
        raw_data = self.selected_data[row_index]
        processed_data = {}
        
        # 엑셀 데이터를 홈택스 필드 형식으로 변환
        for col_idx, (excel_col, hometax_field, formatter) in self.field_mapping.items():
            try:
                # 엑셀에서 가져온 raw_data는 컬럼명으로 접근
                if excel_col in raw_data:
                    raw_value = raw_data[excel_col]
                    processed_data[hometax_field] = formatter(raw_value)
                else:
                    processed_data[hometax_field] = ""
            except Exception as e:
                print(f"데이터 변환 오류 - {excel_col}: {e}")
                processed_data[hometax_field] = ""
        
        # 추가 계산 필드
        try:
            supply = int(processed_data.get('supply_amount', '0'))
            tax = int(processed_data.get('tax_amount', '0'))
            processed_data['total_amount'] = str(supply + tax)
        except:
            processed_data['total_amount'] = "0"
        
        # 원본 행 번호 추가
        processed_data['excel_row'] = raw_data.get('excel_row', 0)
        
        return processed_data
    
    def get_all_processed_data(self):
        """선택된 모든 행의 데이터를 가공하여 반환"""
        processed_list = []
        for i in range(len(self.selected_data)):
            row_data = self.get_processed_row_data(i)
            if row_data:
                processed_list.append(row_data)
        return processed_list

async def process_tax_invoices_with_selected_data(page, processor):
    """선택된 엑셀 데이터를 이용한 세금계산서 처리 - 새로운 순차 처리 방식"""
    print("\n=== 선택된 거래명세표 데이터로 세금계산서 자동 처리 ===")
    
    # 순차 처리 방식 사용
    await process_selected_rows_sequentially(page, processor)

async def process_selected_rows_sequentially(page, processor):
    """선택된 행들을 순차적으로 처리 (거래처별 그룹핑)"""
    print("\n=== 선택된 행들 순차 처리 시작 ===")
    
    groups = processor.group_data_by_business_number()
    if not groups:
        print("처리할 그룹이 없습니다.")
        return
    
    print(f"총 {len(groups)}개 거래처 그룹을 순차 처리합니다.")
    
    processed_count = 0
    
    for group_idx, group_data in enumerate(groups, 1):
        try:
            first_row = group_data[0]
            business_number = str(first_row.get('등록번호', '')).strip()
            company_name = first_row.get('상호', '미상')
            
            print(f"\n[{group_idx}/{len(groups)}] 거래처 그룹 처리 시작")
            print(f"   거래처: {business_number} ({company_name})")
            print(f"   거래건수: {len(group_data)}건")
            
            await process_single_tax_invoice(page, group_data, processor)
            
            processed_count += 1
            
            if group_idx < len(groups):
                await page.wait_for_timeout(2000)
            
        except Exception as e:
            print(f"   ❌ [{group_idx}] 거래처 그룹 처리 중 오류: {e}")
            continue
    
    print(f"\n거래처별 순차 처리 완료!")
    print(f"   처리된 그룹 수: {processed_count} / {len(groups)}")

async def process_single_tax_invoice(page, group_data, processor):
    """월 합계 세금계산서 처리 (16건까지의 거래명세표)"""
    try:
        first_row = group_data[0]
        business_number = str(first_row.get('등록번호', '')).strip()
        
        print(f"      사업자번호 검증 시작: {business_number}")
        
        if not business_number:
            print("❌ 등록번호가 없습니다.")
            for row in group_data:
                processor.write_error_to_excel_q_column(row['excel_row'], "번호없음")
            return

        await input_business_number_and_verify(page, business_number, processor, first_row)
        
    except Exception as e:
        print(f"      사업자번호 검증 처리 실패: {e}")
        business_number = group_data[0].get('등록번호', '알수없음').strip()
        processor.write_error_to_all_matching_business_numbers(business_number, "처리오류")


async def input_transaction_details(page, group_data, processor):
    """거래명세표 입력 (16건까지)"""
    try:
        print(f"      거래명세표 입력: {len(group_data)}건")
        
        # 기본 4건 이외에 추가 품목이 필요한 경우 품목추가 버튼 클릭
        items_to_add = len(group_data) - 4
        if items_to_add > 0:
            print(f"      품목 추가 필요: {items_to_add}건")
            
            for i in range(min(items_to_add, 12)):  # 최대 12번까지 추가 가능
                try:
                    add_btn = page.locator("#mf_txppWframe_btnLsatAddTop")
                    await add_btn.wait_for(state="visible", timeout=3000)
                    await add_btn.click()
                    await page.wait_for_timeout(500)
                    print(f"         품목 {i+1} 추가 완료")
                except Exception as e:
                    print(f"         품목 {i+1} 추가 실패: {e}")
                    break
        
        # 각 거래명세표 행 입력
        for idx, row_data in enumerate(group_data):
            try:
                print(f"      [{idx+1}/{len(group_data)}] 거래명세표 입력 중...")
                
                # 품목별 입력 필드 ID 생성
                row_idx = idx  # 0부터 시작
                
                # 각 필드에 데이터 입력
                await input_transaction_item(page, row_idx, row_data, processor)
                
                print(f"         거래명세표 {idx+1} 입력 완료")
                
            except Exception as e:
                print(f"         거래명세표 {idx+1} 입력 실패: {e}")
                processor.write_error_to_excel(row_data.get('excel_row', 0), "명세표 입력 error")
                continue
        
        print(f"      ✅ 모든 거래명세표 입력 완료: {len(group_data)}건")
        
    except Exception as e:
        print(f"      ❌ 거래명세표 입력 실패: {e}")

async def input_transaction_item(page, row_idx, row_data, processor):
    """개별 거래명세표 행 입력"""
    try:
        # 입력 필드 매핑
        field_mapping = {
            'supply_date': f"#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatSplDdTop",    # 일
            'item_name': f"#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatNmTop",        # 품목
            'spec': f"#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatRszeNmTop",        # 규격
            'quantity': f"#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatQtyTop",       # 수량
            'unit_price': f"#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatUtprcTop",   # 단가
            'supply_amount': f"#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatSplCftTop", # 공급가액
            'tax_amount': f"#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatTxamtTop"    # 세액
        }
        
        # 데이터 매핑 (엑셀 컬럼 → HomeTax 필드)
        input_data = {
            'supply_date': str(row_data.get('작성일자', '')).replace('-', '').replace('/', '').replace('.', ''),  # 엑셀 "작성일자" → HomeTax "공급일자"
            'item_name': str(row_data.get('품명', '')).strip(),        # 엑셀 "품명" → HomeTax "품목"
            'spec': str(row_data.get('규격', '')).strip(),             # 엑셀 "규격" → HomeTax "규격"
            'quantity': str(row_data.get('수량', '')).strip(),         # 엑셀 "수량" → HomeTax "수량"  
            'unit_price': str(row_data.get('단가', '')).strip(),       # 엑셀 "단가" → HomeTax "단가"
            'supply_amount': str(row_data.get('공급가액', '')).strip(), # 엑셀 "공급가액" → HomeTax "공급가액"
            'tax_amount': str(row_data.get('세액', '')).strip()        # 엑셀 "세액" → HomeTax "세액"
        }
        
        # 각 필드에 데이터 입력
        for field_key, selector in field_mapping.items():
            try:
                value = input_data.get(field_key, '')
                if value:
                    input_field = page.locator(selector)
                    await input_field.wait_for(state="visible", timeout=2000)
                    await input_field.clear()
                    await input_field.fill(value)
                    await page.wait_for_timeout(200)
                    print(f"            {field_key}: {value}")
            except Exception as e:
                print(f"            {field_key} 입력 실패: {e}")
                continue
        
    except Exception as e:
        print(f"         개별 거래명세표 입력 실패: {e}")
        processor.write_error_to_excel(row_data.get('excel_row', 0), "개별 입력 error")

async def input_business_number_and_verify(page, business_number, processor, row_data):
    try:
        # 1. 사업자번호 입력
        await page.locator("#mf_txppWframe_edtDmnrBsnoTop").fill(business_number)

        # 6. Alert 감지 리스너를 미리 설정
        dialog_message = None
        dialog_event = asyncio.Event()
        dialog_detected = False

        async def handle_dialog(dialog):
            nonlocal dialog_message, dialog_detected
            print(f"      Alert 감지: {dialog.message}")
            dialog_message = dialog.message
            dialog_detected = True
            await dialog.accept()
            dialog_event.set()

        page.once("dialog", handle_dialog)

        # 2. 확인 버튼 클릭
        await page.locator("#mf_txppWframe_btnDmnrBsnoCnfrTop").click()

        # 잠시 대기하여 반응 확인
        await page.wait_for_timeout(500)

        # 3. 뜨는 화면이 종사업장 메시지 박스인가?
        try:
            branch_popup_close_button = page.locator("#mf_txppWframe_ABTIBsnoUnitPopup2_wframe_btnClose0")
            await branch_popup_close_button.wait_for(state="visible", timeout=1000)
            
            # 4. Yes이면 종사업장 메시지 박스의 닫기
            print("      종사업장 메시지 박스 발견. 닫기 버튼 클릭.")
            await branch_popup_close_button.click()
            
            # 5. 열려있는 세금계산서.xlsx 거래처 시트 q열의 해당거래처의 cell들에 "미등록(주)"라고 기록하고 beep 1회 울림
            print("      엑셀에 '미등록(주)' 기록")
            processor.write_error_to_all_matching_business_numbers(business_number, "미등록(주)")
            await play_beep(1)
            return # 종사업장 처리 후 함수 종료

        except Exception:
            # 종사업장 팝업이 없으면 Alert 대기
            print("      종사업장 메시지 박스 없음. Alert 대기 중...")

            try:
                # Alert가 뜰 때까지 최대 3초 대기
                await asyncio.wait_for(dialog_event.wait(), timeout=3.0)
                # 7. alert이 닫힌 후 1초 대기
                await page.wait_for_timeout(1000)
                print("      Alert 처리 완료")
            except asyncio.TimeoutError:
                print("      Alert가 감지되지 않았습니다.")
                # 리스너 제거
                try:
                    page.remove_listener("dialog", handle_dialog)
                except:
                    pass

            # 8. #mf_txppWframe_edtDmnrTnmNmTop 필드가 active되어 입력 가능한지 체크
            company_name_field = page.locator("#mf_txppWframe_edtDmnrTnmNmTop")
            is_active = False
            try:
                # is_editable() 또는 is_enabled()로 활성 상태 확인
                is_active = await company_name_field.is_editable(timeout=2000)
            except Exception:
                is_active = False

            # 9. active가 되어 있어 입력 가능한가?
            if is_active:
                print("      상호명 필드가 활성화되어 있습니다.")
                company_name = await company_name_field.input_value()
                
                # 9. yes이고 #mf_txppWframe_edtDmnrTnmNmTop 셀렉션 값이 null이면
                if not company_name or company_name.strip() == "":
                    print("      상호명 필드가 비어있습니다. '미등록'으로 기록합니다.")
                    processor.write_error_to_all_matching_business_numbers(business_number, "미등록")
                    await play_beep(2)
                # 10. yes이고 #mf_txppWframe_edtDmnrTnmNmTop 셀렉션 값이 null이 아니면
                else:
                    print(f"      상호명 확인: {company_name}. 거래 내역 입력을 시작합니다.")
                    # 거래 내역 입력 프로세스 호출
                    await process_transaction_details(page, processor, row_data, business_number)
            # 11. #mf_txppWframe_edtDmnrTnmNmTop 필드가 active되어 있지 않아 입력이 가능하지 않은 경우
            else:
                print("      상호명 필드가 활성화되어 있지 않습니다. '번호오류'로 기록합니다.")
                processor.write_error_to_all_matching_business_numbers(business_number, "번호오류")
                await play_beep(3)

    except Exception as e:
        print(f"   ❌ 등록번호 검증 중 심각한 오류 발생: {e}")
        processor.write_error_to_all_matching_business_numbers(business_number, "처리오류")
        await play_beep(3)












async def input_supply_date(page, supply_date):
    """공급일자 입력"""
    try:
        # 날짜 형식 변환
        if isinstance(supply_date, pd.Timestamp):
            supply_date_str = supply_date.strftime("%Y%m%d")
        elif isinstance(supply_date, str):
            # 문자열 날짜를 YYYYMMDD 형식으로 변환
            supply_date_str = supply_date.replace("-", "").replace("/", "").replace(".", "")
        else:
            supply_date_str = str(supply_date)
        
        # 공급일자 입력 필드
        date_input = page.locator("#mf_txppWframe_calWrtDtTop_input")
        await date_input.wait_for(state="visible", timeout=3000)
        await date_input.clear()
        await date_input.fill(supply_date_str)
        print(f"   공급일자 입력 완료: {supply_date_str}")
        
        await page.wait_for_timeout(500)
        
    except Exception as e:
        print(f"   ❌ 공급일자 입력 실패: {e}")

async def auto_process_tax_invoices(page, data_manager):
    """엑셀 데이터를 이용한 세금계산서 자동 처리"""
    try:
        print("세금계산서 자동 처리 시작...")
        
        # 처리할 거래 선택 (최대 3건)
        transactions_to_process = data_manager.transaction_data[:3]
        print(f"처리 예정: {len(transactions_to_process)}건")
        
        for i, transaction in enumerate(transactions_to_process, 1):
            print(f"\n[{i}/{len(transactions_to_process)}] 처리 중: {transaction['상호']}")
            print(f"   품명: {transaction['품명']}")
            print(f"   금액: {transaction['총액']:,}원")
            
            try:
                # 세금계산서 작성 페이지로 이동 (이미 메뉴 네비게이션 완료 상태)
                print("   데이터 입력 대기...")
                await page.wait_for_timeout(2000)
                
                # 실제 입력 필드들을 찾아서 데이터 입력
                await fill_tax_invoice_form(page, transaction)
                
                print(f"   [{i}] {transaction['상호']} 처리 완료")
                await page.wait_for_timeout(3000)  # 다음 처리 전 대기
                
            except Exception as e:
                print(f"   [{i}] 처리 중 오류: {e}")
                continue
        
        print(f"\n세금계산서 자동 처리 완료: {len(transactions_to_process)}건 처리")
        
    except Exception as e:
        print(f"자동 처리 오류: {e}")

async def fill_tax_invoice_form(page, transaction):
    """세금계산서 양식에 데이터 입력"""
    try:
        print(f"      양식 입력 시작...")
        
        # 공통 입력 필드들 (실제 HomeTax 필드명에 맞춰 수정 필요)
        form_data = {
            '거래처명': transaction['상호'],
            '사업자번호': transaction['등록번호'], 
            '품목명': transaction['품명'],
            '규격': transaction['규격'],
            '수량': str(transaction['수량']),
            '단가': str(transaction['단가']),
            '공급가액': str(transaction['공급가액']),
            '세액': str(transaction['세액']),
            '총액': str(transaction['총액'])
        }
        
        # 실제 입력 필드 찾기 및 입력 (예시 - 실제 필드명으로 수정 필요)
        input_selectors = [
            ("상호", "input[name*='상호'], input[id*='company'], input[placeholder*='상호']"),
            ("사업자번호", "input[name*='사업자'], input[id*='business'], input[placeholder*='사업자']"),
            ("품목", "input[name*='품목'], input[id*='item'], input[placeholder*='품목']"),
            ("공급가액", "input[name*='공급'], input[id*='supply'], input[placeholder*='공급']"),
        ]
        
        filled_count = 0
        for field_name, selector in input_selectors:
            try:
                if await page.locator(selector).count() > 0:
                    field_value = form_data.get(field_name, '')
                    if field_value:
                        await page.locator(selector).first.fill(str(field_value))
                        filled_count += 1
                        print(f"         {field_name}: {field_value}")
                        await page.wait_for_timeout(500)
            except Exception as e:
                print(f"         {field_name} 입력 실패: {e}")
                continue
        
        print(f"      양식 입력 완료: {filled_count}개 필드")
        
        # 임시 저장 또는 저장 버튼 클릭 (선택사항)
        save_selectors = [
            "button:has-text('임시저장')",
            "button:has-text('저장')",
            "input[value*='저장']",
            "*[id*='save'], *[id*='temp']"
        ]
        
        for selector in save_selectors:
            try:
                if await page.locator(selector).count() > 0:
                    await page.locator(selector).first.click()
                    print("      임시저장 완료")
                    break
            except:
                continue
                
    except Exception as e:
        print(f"      양식 입력 오류: {e}")

async def hometax_quick_login():
    """
    빠른 홈택스 로그인 자동화 (대기시간 최소화) + 엑셀 데이터 연동
    """
    load_dotenv()
    cert_password = os.getenv("PW")
    if not cert_password:
        print("오류: .env 파일에 PW 변수가 설정되지 않았습니다.")
        return
    
    # 먼저 엑셀 파일 선택 및 행 선택 GUI 실행
    print("=== 엑셀 파일 선택 및 거래명세표 행 선택 ===")
    processor = TaxInvoiceExcelProcessor()
    
    # 엑셀 파일 선택 및 행 선택
    if not processor.select_excel_file_and_process():
        print("엑셀 파일 선택 또는 행 선택이 취소되었습니다.")
        return
    
    print(f"\n선택된 데이터: {len(processor.selected_data)}개 행")
    
    print("\n=== 홈택스 로그인 시작 ===")

    async with async_playwright() as p:
        browser = await p.firefox.launch(
            headless=False, 
            slow_mo=500,
            args=[
                '--disable-web-security',
                '--disable-features=VizDisplayCompositor'
            ]
        )
        
        try:
            page = await browser.new_page()
            page.set_default_timeout(10000)  # 10초로 단축
            
            print("홈택스 페이지 이동...")
            await page.goto("https://hometax.go.kr/websquare/websquare.html?w2xPath=/ui/pp/index_pp.xml&menuCd=index3")
            await page.wait_for_load_state('domcontentloaded')  # networkidle → domcontentloaded로 변경
            
            await page.wait_for_timeout(3000)  # 8초 → 3초로 단축
            
            # 빠른 버튼 찾기 - 직접적인 셀렉터부터 시도
            print("공동·금융인증서 버튼 검색...")
            
            button_selectors = [
                "#mf_txppWframe_loginboxFrame_anchor22",  # 정확한 셀렉터
                "#anchor22",
                "a:has-text('공동인증서')",
                "a:has-text('공동·금융인증서')",
                "a:has-text('금융인증서')"
            ]
            
            login_clicked = False
            for selector in button_selectors:
                try:
                    print(f"시도: {selector}")
                    await page.locator(selector).first.click(timeout=2000)
                    print(f"클릭 성공: {selector}")
                    login_clicked = True
                    break
                except:
                    continue
            
            # iframe 내부에서도 빠르게 시도
            if not login_clicked:
                try:
                    iframe = page.frame_locator("#txppIframe")
                    await iframe.locator("a:has-text('공동')").first.click(timeout=2000)
                    login_clicked = True
                    print("iframe 내부 클릭 성공")
                except:
                    pass
            
            if not login_clicked:
                print("자동 클릭 실패 - 수동으로 '공동·금융인증서' 버튼을 클릭하세요")
                await page.wait_for_timeout(10000)  # 10초만 대기
            
            # #dscert iframe 빠른 대기
            print("인증서 창 대기...")
            dscert_found = False
            
            # 디버깅: 페이지의 현재 상태 확인
            await page.wait_for_timeout(2000)  # 2초 대기 후 상태 확인
            
            # 페이지의 iframe과 popup 요소들 확인
            print("현재 페이지의 iframe 및 popup 요소 검색 중...")
            try:
                # 가능한 인증서 관련 selector들 확인
                possible_selectors = [
                    "#dscert",
                    "iframe[id*='cert']",
                    "iframe[name*='cert']", 
                    "iframe[src*='cert']",
                    "[id*='popup']",
                    "[class*='popup']",
                    "[id*='modal']",
                    "[class*='modal']"
                ]
                
                found_elements = []
                for selector in possible_selectors:
                    try:
                        elements = await page.query_selector_all(selector)
                        if elements:
                            found_elements.append(f"{selector}: {len(elements)}개")
                    except:
                        continue
                
                if found_elements:
                    print(f"발견된 요소들: {', '.join(found_elements)}")
                else:
                    print("인증서 관련 요소를 찾을 수 없습니다.")
                    
            except Exception as e:
                print(f"디버깅 중 오류: {e}")
            
            # iframe 존재 여부와 내용 로딩을 분리하여 처리
            iframe_exists = False
            try:
                # 먼저 iframe이 존재하는지 확인
                await page.wait_for_selector("#dscert", timeout=5000)
                iframe_exists = True
                print("   ✅ #dscert iframe 발견")
            except:
                print("   ❌ #dscert iframe 없음")
            
            if iframe_exists:
                # iframe이 존재하면 내용 로딩을 여러 방법으로 시도
                for i in range(10):
                    try:
                        dscert_iframe = page.frame_locator("#dscert")
                        
                        # 방법 1: body 대기
                        try:
                            await dscert_iframe.locator("body").wait_for(state="visible", timeout=2000)
                            print("인증서 창 발견!")
                            dscert_found = True
                            break
                        except:
                            # 방법 2: 어떤 요소든 로드될 때까지 대기
                            try:
                                await dscert_iframe.locator("*").first.wait_for(state="attached", timeout=2000)
                                print("인증서 창 발견!")
                                dscert_found = True
                                break
                            except:
                                pass
                        
                        print(f"시도 {i+1}/10: iframe 내용 로딩 대기 중...")
                        await page.wait_for_timeout(1500)
                        
                    except Exception as e:
                        print(f"시도 {i+1}/10 실패: {e}")
                        await page.wait_for_timeout(1000)
            
            if not dscert_found:
                print("인증서 창을 찾을 수 없습니다.")
                print("대안 방법: 수동으로 인증서를 선택하신 후 15초 후에 자동으로 계속 진행됩니다.")
                await page.wait_for_timeout(15000)  # 15초 대기
                return
            
            # 인증서 선택 먼저 (Firefox용 최적화)
            print("인증서 선택...")
            try:
                # Firefox에서 더 안정적인 방법으로 인증서 선택
                await page.wait_for_timeout(2000)  # 페이지 안정화 대기
                
                # 강제 클릭 방식 시도 (blockUI 무시)
                cert_selector = dscert_iframe.locator("#row0dataTable > td:nth-child(1) > a").first
                await cert_selector.wait_for(state="attached", timeout=5000)  # visible 대신 attached 사용
                
                # JavaScript로 강제 클릭
                await dscert_iframe.evaluate("""
                    document.querySelector('#row0dataTable > td:nth-child(1) > a').click();
                """)
                print("인증서 선택 완료 (JavaScript 강제 클릭)")
                await page.wait_for_timeout(2000)  # 더 긴 대기 시간
                
            except Exception as e:
                print(f"인증서 선택 실패: {e}")
                # 더 단순한 방법으로 시도
                try:
                    # 테이블의 첫 번째 행 클릭
                    await dscert_iframe.evaluate("""
                        const rows = document.querySelectorAll('#row0dataTable tr');
                        if (rows.length > 0) {
                            rows[0].click();
                        }
                    """)
                    print("대체 방법으로 인증서 선택 완료 (행 클릭)")
                    await page.wait_for_timeout(2000)
                except:
                    print("인증서 선택 실패 - 수동으로 선택하세요")
                    await page.wait_for_timeout(5000)  # 수동 선택 대기
            
            # 비밀번호 빠른 입력
            print("비밀번호 입력...")
            password_input = dscert_iframe.locator("#input_cert_pw").first
            await password_input.wait_for(state="visible", timeout=3000)
            await password_input.fill(cert_password)
            print("비밀번호 입력 완료")
            
            # 확인 버튼 빠른 클릭
            print("확인 버튼 클릭...")
            await page.wait_for_timeout(500)
            
            # 정확한 확인 버튼 셀렉터 사용
            try:
                confirm_btn = dscert_iframe.locator("#btn_confirm_iframe > span").first
                await confirm_btn.wait_for(state="visible", timeout=3000)
                await confirm_btn.click()
                print("확인 버튼 클릭 완료 (정확한 셀렉터)")
            except Exception as e:
                print(f"정확한 셀렉터 실패: {e}")
                # 대체 방법들 시도
                try:
                    confirm_btn = dscert_iframe.locator("#btn_confirm_iframe").first
                    await confirm_btn.click(timeout=3000)
                    print("확인 버튼 클릭 완료 (대체 방법 1)")
                except:
                    try:
                        confirm_btn = dscert_iframe.locator("button:has-text('확인'), input[value*='확인']").first
                        await confirm_btn.click(timeout=3000)
                        print("확인 버튼 클릭 완료 (대체 방법 2)")
                    except:
                        print("확인 버튼 클릭 실패 - 수동으로 클릭하세요")
            
            # 팝업창 및 Alert 처리 (선택적)
            print("팝업창/Alert 확인 중...")
            
            # 현재 URL 저장 (변수 선언)
            current_initial_url = page.url
            
            # Alert 핸들러 미리 등록 (나타나면 자동 처리)
            dialog_handled = False
            def handle_dialog(dialog):
                nonlocal dialog_handled
                dialog_handled = True
                print(f"Alert 감지 및 처리: '{dialog.message}'")
                asyncio.create_task(dialog.accept())
            
            page.on("dialog", handle_dialog)
            
            # 짧은 시간 동안만 팝업/Alert 확인 (최대 3초)
            popup_found = False
            for check in range(3):
                await page.wait_for_timeout(1000)
                
                # 새로운 팝업창 확인 (context를 통해 접근)
                try:
                    context_pages = page.context.pages
                    if len(context_pages) > 1:  # 메인 페이지 외에 다른 페이지가 있는 경우
                        print(f"새 팝업창 감지: {len(context_pages)}개 페이지 중 {len(context_pages) - 1}개 팝업")
                        popup_found = True
                        
                        # 메인 페이지가 아닌 창들 닫기
                        for popup_page in context_pages:
                            if popup_page != page:
                                try:
                                    await popup_page.close()
                                    print("팝업창 닫기 완료")
                                except:
                                    pass
                        break
                except Exception as e:
                    # 팝업창 확인 중 오류가 발생해도 계속 진행
                    pass
                
                # Alert 처리됨 확인
                if dialog_handled:
                    print("Alert 처리 완료")
                    popup_found = True
                    break
                
                # 로그인이 이미 진행되었는지 확인 (URL 변경)
                if page.url != current_initial_url:
                    print("로그인 진행 중 - 팝업 확인 건너뜀")
                    break
            
            if not popup_found and not dialog_handled:
                print("팝업창/Alert 없음 - 정상 진행")
            
            # Alert 핸들러 제거
            page.remove_listener("dialog", handle_dialog)
            
            # 로그인 완료 정확한 확인
            print("로그인 처리 중...")
            final_initial_url = page.url
            
            login_confirmed = False
            for i in range(15):  # 15초까지 확인
                await page.wait_for_timeout(1000)
                current_url = page.url
                current_title = await page.title()
                
                # URL 변경 확인
                if current_url != final_initial_url:
                    print("로그인 성공! URL 변경 감지")
                    print(f"   새 URL: {current_url}")
                    login_confirmed = True
                    break
                
                # 페이지 제목 확인
                if any(keyword in current_title.lower() for keyword in ['main', 'home', '홈', '메인', '국세청']):
                    print(f"로그인 성공! 메인페이지 접근: {current_title}")
                    login_confirmed = True
                    break
                
                # 인증서 창이 사라졌는지 확인 (로그인 성공 신호)
                try:
                    dscert_visible = await page.locator("#dscert").is_visible()
                    if not dscert_visible:
                        print("로그인 성공! 인증서 창 사라짐 확인")
                        login_confirmed = True
                        break
                except:
                    pass
                
                # 로그인 관련 요소 확인
                try:
                    logout_btn = await page.locator("a:has-text('로그아웃'), button:has-text('로그아웃')").count()
                    if logout_btn > 0:
                        print("로그인 성공! 로그아웃 버튼 확인")
                        login_confirmed = True
                        break
                except:
                    pass
                
                if i % 3 == 2:
                    print(f"   대기 중... ({i + 1}/15초)")
            
            # 최종 상태 확인
            final_url = page.url
            final_title = await page.title()
            
            print(f"\n=== 최종 로그인 결과 ===")
            print(f"최종 URL: {final_url}")
            print(f"최종 제목: {final_title}")
            
            if login_confirmed:
                print("홈택스 자동 로그인 성공!")
                
                # Alert창 X버튼으로 닫기
                print("\n=== Alert창 X버튼 닫기 ===")
                try:
                    # 정확한 X버튼 클릭
                    close_button = page.locator("#mf_txppWframe_UTXPPABB29_wframe_btnCloseInvtSpec")
                    await close_button.wait_for(state="visible", timeout=5000)
                    await close_button.click()
                    print("   X버튼으로 알림창 닫기 완료")
                    await page.wait_for_timeout(2000)
                    
                except Exception as e:
                    print(f"X버튼 클릭 실패: {e}")
                    # 대체 방법으로 Alert 처리
                    try:
                        await page.evaluate("""
                            if (window.confirm) window.confirm = function() { return true; };
                            if (window.alert) window.alert = function() { return true; };
                        """)
                        print("   JavaScript Alert 무력화 완료")
                    except:
                        pass
                
                # 추가 메뉴 네비게이션
                print("\n=== 메뉴 네비게이션 시작 ===")
                await page.wait_for_timeout(3000)  # 더 긴 안정화 대기
                
                try:
                    # 1단계: #mf_wfHeader_wq_uuid_333 선택 (Alert창 닫은 후 첫 번째 메뉴)
                    print("1단계: 첫 번째 메뉴 선택 (#mf_wfHeader_wq_uuid_333)...")
                    
                    first_menu_selectors = [
                        "#mf_wfHeader_wq_uuid_333",
                        "*[id*='wq_uuid_333']",
                        "*[id*='wfHeader'] *[id*='333']",
                        "a[href*='333'], button[id*='333']"
                    ]
                    
                    first_clicked = False
                    for selector in first_menu_selectors:
                        try:
                            print(f"   시도: {selector}")
                            first_menu = page.locator(selector).first
                            await first_menu.wait_for(state="visible", timeout=3000)
                            await first_menu.click()
                            print(f"   첫 번째 메뉴 클릭 성공: {selector}")
                            first_clicked = True
                            break
                        except:
                            continue
                    
                    if not first_clicked:
                        print("   첫 번째 메뉴를 찾을 수 없습니다 - 수동으로 선택하세요")
                        await page.wait_for_timeout(10000)  # 수동 선택 대기
                    else:
                        await page.wait_for_timeout(2000)
                    
                    # 2단계: #combineMenuAtag_4601010100 > span 선택 (두 번째 메뉴)
                    print("2단계: 두 번째 메뉴 선택 (#combineMenuAtag_4601010100)...")
                    
                    second_menu_selectors = [
                        "#combineMenuAtag_4601010100 > span",
                        "#combineMenuAtag_4601010100",
                        "*[id*='combineMenu'][id*='4601010100'] > span",
                        "*[id*='combineMenu'][id*='4601010100']",
                        "a[href*='4601010100'] > span",
                        "a[href*='4601010100']"
                    ]
                    
                    second_clicked = False
                    for selector in second_menu_selectors:
                        try:
                            print(f"   시도: {selector}")
                            second_menu = page.locator(selector).first
                            await second_menu.wait_for(state="visible", timeout=3000)
                            await second_menu.click()
                            print(f"   두 번째 메뉴 클릭 성공: {selector}")
                            second_clicked = True
                            break
                        except:
                            continue
                    
                    if not second_clicked:
                        print("   두 번째 메뉴를 찾을 수 없습니다 - 수동으로 선택하세요")
                        await page.wait_for_timeout(10000)  # 수동 선택 대기
                    else:
                        await page.wait_for_timeout(2000)
                    
                    # 2단계 메뉴 클릭 후 팝업 처리
                    print("3단계: 2단계 메뉴 클릭 후 팝업 처리...")
                    await page.wait_for_timeout(2000)  # 팝업이 나타날 시간 대기
                    
                    try:
                        # Alert 대화상자 자동 처리
                        alert_count = 0
                        def handle_second_dialog(dialog):
                            nonlocal alert_count
                            alert_count += 1
                            print(f"   Alert {alert_count} 감지 및 처리: '{dialog.message}'")
                            asyncio.create_task(dialog.accept())
                        
                        page.on("dialog", handle_second_dialog)
                        
                        # 새 팝업창 확인 및 닫기
                        popup_processed = False
                        for check in range(5):  # 5초간 확인
                            await page.wait_for_timeout(1000)
                            
                            # 새로운 팝업창 확인
                            try:
                                context_pages = page.context.pages
                                if len(context_pages) > 1:
                                    print(f"   새 팝업창 감지: {len(context_pages) - 1}개")
                                    
                                    # 메인 페이지가 아닌 모든 창 닫기
                                    for popup_page in context_pages:
                                        if popup_page != page:
                                            try:
                                                await popup_page.close()
                                                print("   새 팝업창 닫기 완료")
                                                popup_processed = True
                                            except:
                                                pass
                            except:
                                pass
                            
                            # Alert 처리됨 확인
                            if alert_count > 0:
                                print(f"   Alert {alert_count}개 처리 완료")
                                popup_processed = True
                        
                        # 알림창 확인 버튼으로 닫기
                        try:
                            print("   알림창 확인 버튼 찾는 중...")
                            notification_confirm = page.locator("#mf_txppWframe_UTEETZZD02_wframe_btnProcess")
                            await notification_confirm.wait_for(state="visible", timeout=3000)
                            await notification_confirm.click()
                            print("   알림창 확인 버튼 클릭 완료")
                            popup_processed = True
                            await page.wait_for_timeout(1000)
                        except Exception as e:
                            print(f"   알림창 확인 버튼 없음 또는 클릭 실패: {e}")
                        
                        # Alert 핸들러 제거
                        page.remove_listener("dialog", handle_second_dialog)
                        
                        if popup_processed:
                            print("   팝업/Alert 처리 완료")
                        else:
                            print("   팝업/Alert 없음 - 정상 진행")
                            
                    except Exception as popup_error:
                        print(f"   팝업 처리 중 오류: {popup_error}")
                    
                    await page.wait_for_timeout(2000)
                    print("전체 메뉴 네비게이션 완료!")
                    
                except Exception as nav_error:
                    print(f"❌ 메뉴 네비게이션 오류: {nav_error}")
                    print("   수동으로 메뉴를 선택해주세요.")
                
            else:
                print("⚠️  로그인 상태 확인 필요")
                print("   브라우저에서 직접 확인해주세요.")
            
            print(f"\n=== 세금계산서 자동 처리 시작 ===")
            
            # 선택된 엑셀 데이터를 이용한 세금계산서 자동화
            print("선택된 거래명세표 데이터로 세금계산서 자동화를 시작합니다.")
            print("1. 등록번호 검증 및 자동 입력")
            print("2. 공급일자 자동 입력")  
            print("3. 에러 처리 (번호 error, 미등록 구분)")
            
            await process_tax_invoices_with_selected_data(page, processor)
            
            print(f"\n=== 처리 완료 - 결과 확인 시간 (15초) ===")
            await page.wait_for_timeout(15000)
            
        except Exception as e:
            print(f"오류: {e}")
        finally:
            await browser.close()

def check_dependencies():
    """필수 패키지 확인 및 설치"""
    required_packages = ['openpyxl', 'psutil', 'xlwings', 'pywin32']
    print("📦 의존성 패키지 확인 중...")
    
    for package in required_packages:
        try:
            __import__(package.replace('pywin32', 'win32gui'))  # pywin32는 win32gui로 import
            print(f"[OK] {package} 설치됨")
        except ImportError:
            print(f"[ERROR] {package} 미설치 - 자동 설치 중...")
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", package])
                print(f"[OK] {package} 설치 완료")
            except subprocess.CalledProcessError as e:
                print(f"[ERROR] {package} 설치 실패: {e}")
                print(f"수동 설치 필요: pip install {package}")

if __name__ == "__main__":
    print("홈택스 세금계산서 자동화 프로그램")
    print("=" * 50)
    
    # 의존성 확인
    check_dependencies()
    
    # 메인 프로그램 실행
    asyncio.run(hometax_quick_login())
