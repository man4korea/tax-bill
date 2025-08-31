# 📁 C:\APP\tax-bill\core\tax-invoice\hometax_tax_invoice.py
# Create at 2508312118 Ver1.00
#-*- coding: utf-8 -*-
import asyncio
import os
import sys
import subprocess
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime
from dotenv import load_dotenv
from playwright.async_api import async_playwright
from excel_data_manager import ExcelDataManager
from hometax_security_manager import HomeTaxSecurityManager
import pandas as pd

# 통합 엑셀 처리 모듈 import
from excel_unified_processor import create_transaction_processor

# 공통 로그인 모듈 import
from hometax_login_module import hometax_login_dispatcher

# 최적화된 모듈들 import
from hometax_utils import (
    play_beep, format_date, format_business_number, format_number,
    FieldCollector, SelectorManager, MenuNavigator, DialogHandler
)
from hometax_transaction_processor import (
    process_transaction_details,
    get_same_business_number_rows,
    check_and_update_supply_date,
    input_transaction_items_basic,
    input_transaction_items_extended,
    input_single_transaction_item,
    finalize_transaction_summary,
    verify_and_calculate_credit,
    handle_issuance_alerts,
    write_to_tax_invoice_sheet,
    clear_form_fields
)

class TaxInvoiceExcelProcessor:
    """ExcelUnifiedProcessor 어댑터 클래스 - 기존 인터페이스 호환성 유지"""
    
    def __init__(self):
        # 통합 프로세서 생성 - 거래명세표 시트용
        self.processor = create_transaction_processor()
        
        # 기존 인터페이스 호환을 위한 속성들 (통합 프로세서에서 위임)
        self.selected_rows = None
        self.selected_data = None
        self.excel_file_path = None
        self.headers = None
        
        # 호환성을 위한 속성 위임 
        self.field_mapping = getattr(self.processor, 'field_mapping', {})
        self.base_selectors = getattr(self.processor, 'base_selectors', {})
        self.item_selectors = getattr(self.processor, 'item_selectors', {})
    
    def _format_date(self, value):
        """날짜 형식 변환"""
        return format_date(value)
    
    def _format_business_number(self, value):
        """사업자번호 형식 변환"""
        return format_business_number(value)
    
    def _format_number(self, value):
        """숫자 형식 변환"""
        return format_number(value) or "0"
    
    def write_error_to_excel(self, row_number, error_message="error"):
        """엑셀 파일의 지정된 행 발행일 열에 에러 메시지 작성"""
        return self.processor.status_recorder.write_status(row_number, error_message)
    
    def write_error_to_excel_q_column(self, row_number, error_message="번호오류"):
        """엑셀 파일의 거래명세표 시트 Q열(발행일)에 에러 메시지 작성 (단일 행)"""
        return self.processor.status_recorder.write_status_to_column(row_number, error_message, 17)  # Q열 = 17번째 컬럼
        
         
    def write_completion_to_excel_q_column(self, row_number, completion_message="완료"):
        """엑셀 파일의 거래명세표 시트 Q열(발행일)에 완료 메시지 작성 (단일 행)"""
        return self.processor.status_recorder.write_status_to_column(row_number, completion_message, 17)  # Q열 = 17번째 컬럼
        
      
    def write_error_to_all_matching_business_numbers(self, business_number, error_message="번호오류"):
        """같은 사업자등록번호를 가진 모든 행의 Q열에 에러 메시지 작성"""
        if not self.excel_file_path:
            print("[ERROR] 엑셀 파일 경로가 없습니다.")
            return False
        
        try:
            import pandas as pd
            
            print(f"같은 등록번호({business_number})를 가진 모든 행에 Q열 에러 기록 중...")
            
            # pandas로 데이터 읽기 (행 찾기용)
            try:
                df = pd.read_excel(self.excel_file_path, sheet_name='거래명세표')
            except:
                df = pd.read_excel(self.excel_file_path)  # 기본 시트 사용
            
            # 사업자등록번호 형식 통일 (하이픈 제거)
            target_business_number = str(business_number).replace('-', '').strip()
            
            # 같은 등록번호를 가진 모든 행 찾기
            matching_rows = []
            for idx, row in df.iterrows():
                row_business_number = str(row.get('등록번호', '')).replace('-', '').strip()
                if row_business_number == target_business_number:
                    excel_row_number = idx + 2  # pandas index는 0부터, 엑셀은 1부터, 헤더 고려하면 +2
                    matching_rows.append(excel_row_number)
            
            if not matching_rows:
                print(f"[ERROR] 등록번호 {business_number}와 일치하는 행을 찾을 수 없습니다.")
                return False
            
            print(f"발견된 일치 행들: {matching_rows} (총 {len(matching_rows)}개)")
            
            # 방법 1: xlwings를 사용해서 열린 엑셀 파일에 직접 쓰기 시도
            try:
                import xlwings as xw
                
                # 현재 열려있는 엑셀 앱에 연결
                try:
                    app = xw.apps.active
                except:
                    app = xw.App(visible=True, add_book=False)
                
                # 열린 워크북 찾기
                workbook_name = self.excel_file_path.split("\\")[-1]  # 파일명만 추출
                wb = None
                
                for book in app.books:
                    if book.name == workbook_name:
                        wb = book
                        break
                
                if wb:
                    # "거래명세표" 시트 선택
                    ws = None
                    for sheet in wb.sheets:
                        if sheet.name == "거래명세표":
                            ws = sheet
                            break
                    
                    if not ws:
                        ws = wb.sheets[0]  # 첫 번째 시트 사용
                    
                    # Q열(17번째 열)에 에러 메시지 기록
                    updated_count = 0
                    for row_number in matching_rows:
                        try:
                            ws.range(f'Q{row_number}').value = error_message
                            updated_count += 1
                            print(f"   행 {row_number} Q열에 '{error_message}' 작성 완료 (xlwings)")
                        except Exception as e:
                            print(f"   행 {row_number} Q열 작성 실패 (xlwings): {e}")
                            continue
                    
                    # 저장
                    wb.save()
                    
                    print(f"[OK] 등록번호 {business_number}의 모든 행 Q열 에러 기록 완료 (xlwings): {updated_count}/{len(matching_rows)}개 행")
                    return True
                    
            except ImportError:
                print("   xlwings가 설치되지 않았습니다. openpyxl 방법을 시도합니다...")
            except Exception as e:
                print(f"   xlwings 방법 실패: {e}")
            
            # 방법 2: openpyxl로 파일 직접 수정 (엑셀이 닫혀있을 때만 가능)
            from openpyxl import load_workbook
            
            workbook = load_workbook(self.excel_file_path)
            
            # 거래명세표 시트 선택
            if "거래명세표" in workbook.sheetnames:
                worksheet = workbook["거래명세표"]
            else:
                worksheet = workbook.active
                print(f"경고: '거래명세표' 시트를 찾을 수 없어 기본 시트({worksheet.title}) 사용")
            
            # 모든 일치하는 행의 Q열(17번째 열)에 에러 메시지 작성
            updated_count = 0
            for row_number in matching_rows:
                try:
                    worksheet.cell(row=row_number, column=17, value=error_message)
                    updated_count += 1
                    print(f"   행 {row_number} Q열에 '{error_message}' 작성 완료 (openpyxl)")
                except Exception as e:
                    print(f"   행 {row_number} Q열 작성 실패 (openpyxl): {e}")
                    continue
            
            # 파일 저장
            workbook.save(self.excel_file_path)
            workbook.close()
            
            print(f"[OK] 등록번호 {business_number}의 모든 행 Q열 에러 기록 완료 (openpyxl): {updated_count}/{len(matching_rows)}개 행")
            return True
            
        except PermissionError as pe:
            print(f"[ERROR] 파일 권한 오류: {pe}")
            print("   [FIX] 해결 방법:")
            print("   1. 엑셀 파일이 열려있다면 파일을 닫고 다시 시도하세요")
            print("   2. 또는 xlwings를 설치하세요: pip install xlwings")
            return False
            
        except Exception as e:
            print(f"[ERROR] 같은 등록번호 모든 행 Q열 에러 기록 실패: {e}")
            return False
    
    def write_tax_invoice_data(self, tax_invoice_data):
        """세금계산서 시트에 데이터 기록"""
        if not self.excel_file_path:
            print("[ERROR] 엑셀 파일 경로가 없습니다.")
            return False
        
        try:
            from openpyxl import load_workbook
            
            print(f"세금계산서 시트에 데이터 기록 중...")
            
            # 방법 1: xlwings를 사용해서 열린 엑셀 파일에 직접 쓰기 시도
            try:
                import xlwings as xw
                
                # 현재 열려있는 엑셀 앱에 연결
                try:
                    app = xw.apps.active
                except:
                    app = xw.App(visible=True, add_book=False)
                
                # 열린 워크북 찾기
                workbook_name = self.excel_file_path.split("\\")[-1]
                wb = None
                
                for book in app.books:
                    if book.name == workbook_name:
                        wb = book
                        break
                
                if wb:
                    # "세금계산서" 시트 찾기 또는 생성
                    ws = None
                    for sheet in wb.sheets:
                        if sheet.name == "세금계산서":
                            ws = sheet
                            break
                    
                    if not ws:
                        # 세금계산서 시트가 없으면 생성
                        ws = wb.sheets.add("세금계산서")
                        # 헤더 작성
                        headers = ['공급일자', '등록번호', '상호', '이메일', '', '품목', '규격', '수량', '공급가액', '세액', '합계금액', '기간및건수']
                        for i, header in enumerate(headers, 1):
                            ws.range(f'{chr(64+i)}1').value = header
                    
                    # 마지막 행 찾기
                    last_row = 1
                    while ws.range(f'A{last_row}').value is not None:
                        last_row += 1
                    
                    # 데이터 기록
                    for col_letter, value in tax_invoice_data.items():
                        if value:  # 값이 있을 때만 기록
                            ws.range(f'{col_letter.upper()}{last_row}').value = value
                    
                    # 저장
                    wb.save()
                    
                    print(f"[OK] 세금계산서 시트에 데이터 기록 완료 (xlwings): 행 {last_row}")
                    return True
                    
            except ImportError:
                print("   xlwings가 설치되지 않았습니다. openpyxl 방법을 시도합니다...")
            except Exception as e:
                print(f"   xlwings 방법 실패: {e}")
            
            # 방법 2: openpyxl로 파일 직접 수정
            workbook = load_workbook(self.excel_file_path)
            
            # 세금계산서 시트 찾기 또는 생성
            if "세금계산서" in workbook.sheetnames:
                worksheet = workbook["세금계산서"]
            else:
                worksheet = workbook.create_sheet("세금계산서")
                # 헤더 작성
                headers = ['공급일자', '등록번호', '상호', '이메일', '', '품목', '규격', '수량', '공급가액', '세액', '합계금액', '기간및건수']
                for i, header in enumerate(headers, 1):
                    worksheet.cell(row=1, column=i, value=header)
            
            # 마지막 행 찾기
            last_row = 1
            while worksheet.cell(row=last_row, column=1).value is not None:
                last_row += 1
            
            # 컬럼 매핑 (a=1, b=2, c=3, ...)
            column_mapping = {
                'a': 1, 'b': 2, 'c': 3, 'd': 4, 'e': 5, 'f': 6, 'g': 7, 'h': 8,
                'i': 9, 'j': 10, 'k': 11, 'l': 12
            }
            
            # 데이터 기록
            for col_letter, value in tax_invoice_data.items():
                if value and col_letter.lower() in column_mapping:
                    col_num = column_mapping[col_letter.lower()]
                    worksheet.cell(row=last_row, column=col_num, value=value)
            
            # 파일 저장
            workbook.save(self.excel_file_path)
            workbook.close()
            
            print(f"[OK] 세금계산서 시트에 데이터 기록 완료 (openpyxl): 행 {last_row}")
            return True
            
        except PermissionError as pe:
            print(f"[ERROR] 파일 권한 오류: {pe}")
            print("   [FIX] 해결 방법:")
            print("   1. 엑셀 파일이 열려있다면 파일을 닫고 다시 시도하세요")
            print("   2. 또는 xlwings를 설치하세요: pip install xlwings")
            return False
            
        except Exception as e:
            print(f"[ERROR] 세금계산서 시트 기록 실패: {e}")
            return False
    
    def check_and_open_excel_file(self):
        """세금계산서.xlsx 파일 체크 및 자동 열기 - 통합 프로세서로 위임"""
        return self.processor.file_manager.check_and_open_file()
        target_filename = "세금계산서.xlsx"
        
  
    def select_excel_file_and_process(self):
        """엑셀 파일 체크/열기 및 거래명세표 시트에서 행 선택 처리 - 통합 프로세서로 위임"""
        result = self.processor.select_file_and_process()
        if result:
            self.selected_rows = result.get('selected_rows')
            self.selected_data = result.get('selected_data')
            self.excel_file_path = result.get('excel_file_path')
            self.headers = result.get('headers')
        return result
       
    
    def parse_row_selection(self, selection, silent=False):
        """행 선택 문자열을 파싱하여 행 번호 리스트 반환 (hometax_excel_integration.py 방식)"""
        if not selection.strip():
            return []
        
        rows = []
        parts = selection.replace(" ", "").split(",")
        
        for part in parts:
            if "-" in part:
                # 범위 처리 (예: 2-5)
                try:
                    start, end = part.split("-", 1)
                    start_row = int(start)
                    end_row = int(end)
                    if start_row <= end_row:
                        rows.extend(range(start_row, end_row + 1))
                    else:
                        if not silent:
                            print(f"[ERROR] 잘못된 범위: {part}")
                except ValueError:
                    if not silent:
                        print(f"[ERROR] 잘못된 범위: {part}")
            else:
                # 단일 행 처리
                try:
                    row = int(part)
                    if row > 1:  # 헤더 행 제외
                        rows.append(row)
                    else:
                        if not silent:
                            print(f"[ERROR] 잘못된 행 번호: {part}")
                except ValueError:
                    if not silent:
                        print(f"[ERROR] 잘못된 행 번호: {part}")
        
        return sorted(set(rows))  # 중복 제거 및 정렬
    
    def show_row_selection_gui(self):
        """행 선택 GUI 표시 - 통합 프로세서로 위임"""
        result = self.processor.row_selector.show_gui()
        if result:
            self.selected_rows = result.get('selected_rows')
            self.selected_data = result.get('selected_data')
            self.excel_file_path = result.get('excel_file_path') 
            self.headers = result.get('headers')
        return result

    
    def group_data_by_business_number(self):
        """사업자번호별로 월 합계 세금계산서 그룹핑 (16건씩) - 통합 프로세서로 위임"""
        return self.processor.data_processor.group_by_business_number()
        
    def get_processed_row_data(self, row_index):
        """선택된 행의 데이터를 홈택스 필드용으로 가공하여 반환 - 통합 프로세서로 위임"""
        return self.processor.data_processor.get_processed_row_data(row_index)
    
    
    def get_all_processed_data(self):
        """선택된 모든 행의 데이터를 가공하여 반환 - 통합 프로세서로 위임"""
        return self.processor.data_processor.get_all_processed_data()

async def process_tax_invoices_with_selected_data(page, processor):
    """선택된 엑셀 데이터를 이용한 세금계산서 처리 - 새로운 순차 처리 방식"""
    print("\n=== 선택된 거래명세표 데이터로 세금계산서 자동 처리 ===")
    
    # 순차 처리 방식 사용
    await process_selected_rows_sequentially(page, processor)

async def process_selected_rows_sequentially(page, processor):
    """선택된 행들을 순차적으로 처리 (거래처별 그룹핑)"""
    print("\n=== 선택된 행들 순차 처리 시작 ===")
    
    groups = processor.group_data_by_business_number()
    if not groups:
        print("처리할 그룹이 없습니다.")
        return
    
    print(f"총 {len(groups)}개 거래처 그룹을 순차 처리합니다.")
    
    processed_count = 0
    
    for group_idx, group_data in enumerate(groups, 1):
        try:
            first_row = group_data[0]
            business_number = str(first_row.get('등록번호', '')).strip()
            company_name = first_row.get('상호', '미상')
            
            print(f"\n[{group_idx}/{len(groups)}] 거래처 그룹 처리 시작")
            print(f"   거래처: {business_number} ({company_name})")
            print(f"   거래건수: {len(group_data)}건")
            
            await process_single_tax_invoice(page, group_data, processor)
            
            processed_count += 1
            
            if group_idx < len(groups):
                await page.wait_for_timeout(2000)
            
        except Exception as e:
            print(f"   [ERROR] [{group_idx}] 거래처 그룹 처리 중 오류: {e}")
            continue
    
    print(f"\n거래처별 순차 처리 완료!")
    print(f"   처리된 그룹 수: {processed_count} / {len(groups)}")
    
    # 모든 거래처 처리 완료 후 로그아웃
    try:
        print("\n[LOGOUT] 모든 작업 완료 - 로그아웃 처리 중...")
        await page.wait_for_timeout(2000)  # 안정화 대기
        
        # 로그아웃 버튼 클릭
        logout_btn = page.locator("#mf_wfHeader_group1503")
        await logout_btn.wait_for(state="visible", timeout=5000)
        await logout_btn.click()
        print("[OK] 로그아웃 버튼 클릭 완료")
        
        # 로그아웃 확인 대기
        await page.wait_for_timeout(3000)
        print("[OK] 로그아웃 처리 완료")
        
    except Exception as logout_error:
        print(f"[WARN] 로그아웃 처리 실패 (무시하고 계속): {logout_error}")

async def process_single_tax_invoice(page, group_data, processor):
    """월 합계 세금계산서 처리 (16건까지의 거래명세표)"""
    try:
        first_row = group_data[0]
        business_number = str(first_row.get('등록번호', '')).strip()
        
        print(f"      사업자번호 검증 시작: {business_number}")
        
        if not business_number:
            print("[ERROR] 등록번호가 없습니다.")
            for row in group_data:
                processor.write_error_to_excel_q_column(row['excel_row'], "번호없음")
            return

        # 사업자번호 검증
        await input_business_number_and_verify(page, business_number, processor, first_row)
        
        # 거래명세표 입력
        await input_transaction_details(page, group_data, processor)
        
        # 발급보류 처리
        from hometax_transaction_processor import finalize_transaction_summary, write_to_tax_invoice_sheet
        
        issuance_success = await finalize_transaction_summary(page, group_data, processor, business_number)
        
        # 세금계산서 시트에 기록
        if issuance_success:
            await write_to_tax_invoice_sheet(page, processor, group_data, business_number)
            print(f"      [OK] 세금계산서 처리 완료: {business_number}")
        else:
            print(f"      [ERROR] 발급보류 실패: {business_number}")
            processor.write_error_to_all_matching_business_numbers(business_number, "발급실패")
        
    except Exception as e:
        print(f"      사업자번호 검증 처리 실패: {e}")
        business_number = group_data[0].get('등록번호', '알수없음').strip()
        processor.write_error_to_all_matching_business_numbers(business_number, "처리오류")


async def input_transaction_details(page, group_data, processor):
    """거래명세표 입력 (16건까지)"""
    try:
        print(f"      거래명세표 입력: {len(group_data)}건")
        
        # 기본 4건 이외에 추가 품목이 필요한 경우 품목추가 버튼 클릭
        items_to_add = len(group_data) - 4
        if items_to_add > 0:
            print(f"      품목 추가 필요: {items_to_add}건")
            
            for i in range(min(items_to_add, 12)):  # 최대 12번까지 추가 가능
                try:
                    add_btn = page.locator("#mf_txppWframe_btnLsatAddTop")
                    await add_btn.wait_for(state="visible", timeout=3000)
                    await add_btn.click()
                    await page.wait_for_timeout(500)
                    print(f"         품목 {i+1} 추가 완료")
                except Exception as e:
                    print(f"         품목 {i+1} 추가 실패: {e}")
                    break
        
        # 각 거래명세표 행 입력
        for idx, row_data in enumerate(group_data):
            try:
                print(f"      [{idx+1}/{len(group_data)}] 거래명세표 입력 중...")
                
                # 품목별 입력 필드 ID 생성
                row_idx = idx  # 0부터 시작
                
                # 각 필드에 데이터 입력
                await input_transaction_item(page, row_idx, row_data, processor)
                
                print(f"         거래명세표 {idx+1} 입력 완료")
                
            except Exception as e:
                print(f"         거래명세표 {idx+1} 입력 실패: {e}")
                processor.write_error_to_excel(row_data.get('excel_row', 0), "명세표 입력 error")
                continue
        
        print(f"      [OK] 모든 거래명세표 입력 완료: {len(group_data)}건")
        
    except Exception as e:
        print(f"      [ERROR] 거래명세표 입력 실패: {e}")

async def input_transaction_item(page, row_idx, row_data, processor):
    """개별 거래명세표 행 입력"""
    try:
        print(f"            엑셀 데이터 키들: {list(row_data.keys())}")  # 디버깅용
        # 입력 필드 매핑
        field_mapping = {
            'supply_date': f"#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatSplDdTop",    # 일
            'item_name': f"#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatNmTop",        # 품목
            'spec': f"#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatRszeNmTop",        # 규격
            'quantity': f"#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatQtyTop",       # 수량
            'unit_price': f"#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatUtprcTop",   # 단가
            'supply_amount': f"#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatSplCftTop", # 공급가액
            'tax_amount': f"#mf_txppWframe_genEtxivLsatTop_{row_idx}_edtLsatTxamtTop"    # 세액
        }
        
        # 유연한 컬럼명 매핑을 위한 헬퍼 함수
        def get_field_value(row_data, field_names, is_date=False, field_type=""):
            """여러 가능한 컬럼명에서 값을 찾아 반환"""
            for field_name in field_names:
                if field_name in row_data and row_data[field_name]:
                    value = str(row_data[field_name]).strip()
                    if is_date and value:
                        # 날짜 형식 정리 (YYYY-MM-DD, YYYY/MM/DD 등을 YYYYMMDD로 변환)
                        value = value.replace('-', '').replace('/', '').replace('.', '')
                        # 날짜에서 일자만 추출 (마지막 2자리)
                        if len(value) >= 2:
                            try:
                                import pandas as pd
                                date_obj = pd.to_datetime(row_data[field_name])
                                value = str(date_obj.day)  # 일자만 추출
                                print(f"            {field_type} 매핑: '{field_name}' = {row_data[field_name]} → 일자 {value}")
                            except:
                                value = value[-2:]  # 마지막 2자리만
                                print(f"            {field_type} 매핑: '{field_name}' = {row_data[field_name]} → 일자 {value} (fallback)")
                    else:
                        print(f"            {field_type} 매핑: '{field_name}' = {value}")
                    return value
            print(f"            {field_type} 매핑: 해당 컬럼 없음 (시도한 컬럼들: {field_names})")
            return ''
        
        # 데이터 매핑 (엑셀 컬럼 → HomeTax 필드) - 사용자 엑셀 컬럼명에 맞게 우선순위 조정
        input_data = {
            'supply_date': get_field_value(row_data, ['공급일자', '작성일자', '일자', '날짜', 'supply_date', 'date', 'supply_dt'], is_date=True, field_type="일자"),
            'item_name': get_field_value(row_data, ['품목명', '품명', '품목', 'item_name', 'item', 'product', 'product_name', '상품명', 'name'], field_type="품목"),
            'spec': get_field_value(row_data, ['규격', 'spec', 'specification', 'size'], field_type="규격"),
            'quantity': get_field_value(row_data, ['수량', 'quantity', 'qty', 'amount'], field_type="수량"),
            'unit_price': get_field_value(row_data, ['단가', 'unit_price', 'price', 'unitprice'], field_type="단가"),
            'supply_amount': get_field_value(row_data, ['공급가액', 'supply_amount', 'amount', 'total'], field_type="공급가액"),
            'tax_amount': get_field_value(row_data, ['세액', 'tax_amount', 'tax', 'vat'], field_type="세액")
        }
        
        # 각 필드에 데이터 입력
        for field_key, selector in field_mapping.items():
            try:
                value = input_data.get(field_key, '')
                if value:
                    input_field = page.locator(selector)
                    await input_field.wait_for(state="visible", timeout=2000)
                    await input_field.clear()
                    await input_field.fill(value)
                    await page.wait_for_timeout(200)
                    print(f"            {field_key}: {value}")
                else:
                    print(f"            {field_key}: (빈 값 - 건너뜀)")
            except Exception as e:
                print(f"            {field_key} 입력 실패: {e} (선택자: {selector})")
                continue
        
    except Exception as e:
        print(f"         개별 거래명세표 입력 실패: {e}")
        processor.write_error_to_excel(row_data.get('excel_row', 0), "개별 입력 error")

async def input_business_number_and_verify(page, business_number, processor, row_data):
    try:
        # 1. 사업자번호 입력
        await page.locator("#mf_txppWframe_edtDmnrBsnoTop").fill(business_number)

        # 6. Alert 감지 리스너를 미리 설정
        dialog_message = None
        dialog_event = asyncio.Event()
        dialog_detected = False

        async def handle_dialog(dialog):
            nonlocal dialog_message, dialog_detected
            print(f"      Alert 감지: {dialog.message}")
            dialog_message = dialog.message
            dialog_detected = True
            await dialog.accept()
            dialog_event.set()

        page.once("dialog", handle_dialog)

        # 2. 확인 버튼 클릭 (타임아웃 방지를 위한 다단계 대기)
        confirm_btn = page.locator("#mf_txppWframe_btnDmnrBsnoCnfrTop")
        try:
            # 버튼이 존재하고 클릭 가능할 때까지 대기 (더 긴 타임아웃)
            await confirm_btn.wait_for(state="attached", timeout=15000)
            await confirm_btn.wait_for(state="visible", timeout=5000)
            await page.wait_for_timeout(1000)  # 추가 안정화 대기
            await confirm_btn.click(timeout=10000)
            print(f"      [OK] 사업자번호 확인 버튼 클릭 완료")
        except Exception as click_error:
            print(f"      [ERROR] 확인 버튼 클릭 실패 - 재시도: {click_error}")
            # 재시도 로직 - 페이지 새로고침 후 다시 시도
            await page.reload()
            await page.wait_for_timeout(3000)
            await page.locator("#mf_txppWframe_edtDmnrBsnoTop").fill(business_number)
            await page.wait_for_timeout(1000)
            await confirm_btn.click()

        # 잠시 대기하여 반응 확인
        await page.wait_for_timeout(500)

        # 3. Alert가 이미 감지되었는지 확인
        if dialog_detected:
            print("      Alert 이미 처리됨 - 종사업장 체크 건너뜀")
            await page.wait_for_timeout(1000)
            
            # Alert 메시지에 따른 처리 분기
            if dialog_message and "사업자등록번호를 입력하세요" in dialog_message:
                print(f"      [ERROR] 사업자번호 입력 오류: {dialog_message}")
                processor.write_error_to_all_matching_business_numbers(business_number, "번호오류")
                raise Exception(f"사업자번호 입력 오류: {dialog_message}")
            elif dialog_message and "정상적인 사업자번호" in dialog_message:
                # 사업자번호 검증 완료 후 거래처 정보 수집
                await collect_partner_info_after_verification(page, business_number, processor)
                print("      사업자번호 검증 완료")
                return
            else:
                print(f"      [WARN] 알 수 없는 Alert: {dialog_message}")
                return
        
        # 4. Alert가 없으면 종사업장 메시지 박스 확인
        try:
            branch_popup_close_button = page.locator("#mf_txppWframe_ABTIBsnoUnitPopup2_wframe_btnClose0")
            await branch_popup_close_button.wait_for(state="visible", timeout=1000)
            
            # 5. Yes이면 종사업장 메시지 박스의 닫기
            print("      종사업장 메시지 박스 발견. 닫기 버튼 클릭.")
            await branch_popup_close_button.click()
            
            # 6. 열려있는 세금계산서.xlsx 거래처 시트 q열의 해당거래처의 cell들에 "미등록(주)"라고 기록하고 beep 1회 울림
            print("      엑셀에 '미등록(주)' 기록")
            processor.write_error_to_all_matching_business_numbers(business_number, "미등록(주)")
            await play_beep(1)
            return # 종사업장 처리 후 함수 종료

        except Exception:
            # 종사업장 팝업이 없으면 Alert 대기
            print("      종사업장 메시지 박스 없음. Alert 대기 중...")

            try:
                # Alert가 뜰 때까지 최대 3초 대기
                await asyncio.wait_for(dialog_event.wait(), timeout=3.0)
                # 7. alert이 닫힌 후 1초 대기
                await page.wait_for_timeout(1000)
                
                # Alert 메시지에 따른 처리 분기
                if dialog_message and "사업자등록번호를 입력하세요" in dialog_message:
                    print(f"      [ERROR] 사업자번호 입력 오류: {dialog_message}")
                    processor.write_error_to_all_matching_business_numbers(business_number, "번호오류")
                    raise Exception(f"사업자번호 입력 오류: {dialog_message}")
                elif dialog_message and "정상적인 사업자번호" in dialog_message:
                    # Alert 처리 완료 후 거래처 정보 수집
                    await collect_partner_info_after_verification(page, business_number, processor)
                    print("      Alert 처리 완료")
                else:
                    print(f"      [WARN] 알 수 없는 Alert: {dialog_message}")
            except asyncio.TimeoutError:
                print("      Alert가 감지되지 않았습니다.")
                # 리스너 제거
                try:
                    page.remove_listener("dialog", handle_dialog)
                except:
                    pass

            # 8. #mf_txppWframe_edtDmnrTnmNmTop 필드가 active되어 입력 가능한지 체크
            company_name_field = page.locator("#mf_txppWframe_edtDmnrTnmNmTop")
            is_active = False
            try:
                # is_editable() 또는 is_enabled()로 활성 상태 확인
                is_active = await company_name_field.is_editable(timeout=2000)
            except Exception:
                is_active = False

            # 9. active가 되어 있어 입력 가능한가?
            if is_active:
                print("      상호명 필드가 활성화되어 있습니다.")
                company_name = await company_name_field.input_value()
                
                # 9. yes이고 #mf_txppWframe_edtDmnrTnmNmTop 셀렉션 값이 null이면
                if not company_name or company_name.strip() == "":
                    print("      상호명 필드가 비어있습니다. '미등록'으로 기록합니다.")
                    processor.write_error_to_all_matching_business_numbers(business_number, "미등록")
                    await play_beep(2)
                # 10. yes이고 #mf_txppWframe_edtDmnrTnmNmTop 셀렉션 값이 null이 아니면
                else:
                    print(f"      상호명 확인: {company_name}. 거래 내역 입력을 시작합니다.")
                    # 거래 내역 입력 프로세스 호출
                    await process_transaction_details(page, processor, row_data, business_number)
            # 11. #mf_txppWframe_edtDmnrTnmNmTop 필드가 active되어 있지 않아 입력이 가능하지 않은 경우
            else:
                print("      상호명 필드가 활성화되어 있지 않습니다. '번호오류'로 기록합니다.")
                processor.write_error_to_all_matching_business_numbers(business_number, "번호오류")
                await play_beep(3)

    except Exception as e:
        print(f"   [ERROR] 등록번호 검증 중 심각한 오류 발생: {e}")
        processor.write_error_to_all_matching_business_numbers(business_number, "처리오류")
        await play_beep(3)


async def input_supply_date(page, supply_date):
    """공급일자 입력"""
    try:
        # 날짜 형식 변환
        if isinstance(supply_date, pd.Timestamp):
            supply_date_str = supply_date.strftime("%Y%m%d")
        elif isinstance(supply_date, str):
            # 문자열 날짜를 YYYYMMDD 형식으로 변환
            supply_date_str = supply_date.replace("-", "").replace("/", "").replace(".", "")
        else:
            supply_date_str = str(supply_date)
        
        # 공급일자 입력 필드
        date_input = page.locator("#mf_txppWframe_calWrtDtTop_input")
        await date_input.wait_for(state="visible", timeout=3000)
        await date_input.clear()
        await date_input.fill(supply_date_str)
        print(f"   공급일자 입력 완료: {supply_date_str}")
        
        await page.wait_for_timeout(500)
        
    except Exception as e:
        print(f"   [ERROR] 공급일자 입력 실패: {e}")

async def auto_process_tax_invoices(page, data_manager):
    """엑셀 데이터를 이용한 세금계산서 자동 처리"""
    try:
        print("세금계산서 자동 처리 시작...")
        
        # 처리할 거래 선택 (최대 3건)
        transactions_to_process = data_manager.transaction_data[:3]
        print(f"처리 예정: {len(transactions_to_process)}건")
        
        for i, transaction in enumerate(transactions_to_process, 1):
            print(f"\n[{i}/{len(transactions_to_process)}] 처리 중: {transaction['상호']}")
            print(f"   품명: {transaction['품명']}")
            print(f"   금액: {transaction['총액']:,}원")
            
            try:
                # 세금계산서 작성 페이지로 이동 (이미 메뉴 네비게이션 완료 상태)
                print("   데이터 입력 대기...")
                await page.wait_for_timeout(2000)
                
                # 실제 입력 필드들을 찾아서 데이터 입력
                await fill_tax_invoice_form(page, transaction)
                
                print(f"   [{i}] {transaction['상호']} 처리 완료")
                await page.wait_for_timeout(3000)  # 다음 처리 전 대기
                
            except Exception as e:
                print(f"   [{i}] 처리 중 오류: {e}")
                continue
        
        print(f"\n세금계산서 자동 처리 완료: {len(transactions_to_process)}건 처리")
        
    except Exception as e:
        print(f"자동 처리 오류: {e}")

async def fill_tax_invoice_form(page, transaction):
    """세금계산서 양식에 데이터 입력"""
    try:
        print(f"      양식 입력 시작...")
        
        # 공통 입력 필드들 (실제 HomeTax 필드명에 맞춰 수정 필요)
        form_data = {
            '거래처명': transaction['상호'],
            '사업자번호': transaction['등록번호'], 
            '품목명': transaction['품명'],
            '규격': transaction['규격'],
            '수량': str(transaction['수량']),
            '단가': str(transaction['단가']),
            '공급가액': str(transaction['공급가액']),
            '세액': str(transaction['세액']),
            '총액': str(transaction['총액'])
        }
        
        # 실제 입력 필드 찾기 및 입력 (예시 - 실제 필드명으로 수정 필요)
        input_selectors = [
            ("상호", "input[name*='상호'], input[id*='company'], input[placeholder*='상호']"),
            ("사업자번호", "input[name*='사업자'], input[id*='business'], input[placeholder*='사업자']"),
            ("품목", "input[name*='품목'], input[id*='item'], input[placeholder*='품목']"),
            ("공급가액", "input[name*='공급'], input[id*='supply'], input[placeholder*='공급']"),
        ]
        
        filled_count = 0
        for field_name, selector in input_selectors:
            try:
                if await page.locator(selector).count() > 0:
                    field_value = form_data.get(field_name, '')
                    if field_value:
                        await page.locator(selector).first.fill(str(field_value))
                        filled_count += 1
                        print(f"         {field_name}: {field_value}")
                        await page.wait_for_timeout(500)
            except Exception as e:
                print(f"         {field_name} 입력 실패: {e}")
                continue
        
        print(f"      양식 입력 완료: {filled_count}개 필드")
        
        # 임시 저장 또는 저장 버튼 클릭 (선택사항)
        save_selectors = [
            "button:has-text('임시저장')",
            "button:has-text('저장')",
            "input[value*='저장']",
            "*[id*='save'], *[id*='temp']"
        ]
        
        for selector in save_selectors:
            try:
                if await page.locator(selector).count() > 0:
                    await page.locator(selector).first.click()
                    print("      임시저장 완료")
                    break
            except:
                continue
                
    except Exception as e:
        print(f"      양식 입력 오류: {e}")

async def hometax_tax_invoice_after_login(page, browser):
    """로그인 완료 후 세금계산서 처리 콜백 함수"""
    print("✅ 로그인 완료 - 세금계산서 처리 시작")
    
    # 먼저 엑셀 파일 선택 및 행 선택 GUI 실행
    print("=== 엑셀 파일 선택 및 거래명세표 행 선택 ===")
    processor = TaxInvoiceExcelProcessor()
    
    # 엑셀 파일 선택 및 행 선택
    if not processor.select_excel_file_and_process():
        print("엑셀 파일 선택 또는 행 선택이 취소되었습니다.")
        return None, None
    
    print(f"\n선택된 데이터: {len(processor.selected_data)}개 행")
    
    # 세금계산서 처리 실행
    await process_tax_invoices_with_selected_data(page, processor)
    
    return page, browser


async def hometax_quick_login():
    """
    빠른 홈택스 로그인 자동화 + 세금계산서 처리 (공통 로그인 모듈 사용)
    """
    print("=== 홈택스 세금계산서 자동화 프로그램 ===")
    
    # 공통 로그인 모듈 사용
    result = await hometax_login_dispatcher(hometax_tax_invoice_after_login)
    
    if result:
        print("✅ 세금계산서 자동화 프로세스 완료!")
    else:
        print("❌ 세금계산서 자동화 프로세스 실패")


def check_dependencies():
    """필수 패키지 확인 및 설치"""
    required_packages = ['openpyxl', 'psutil', 'xlwings', 'pywin32']
    print("[INFO] 의존성 패키지 확인 중...")
    
    for package in required_packages:
        try:
            __import__(package.replace('pywin32', 'win32gui'))  # pywin32는 win32gui로 import
            print(f"[OK] {package} 설치됨")
        except ImportError:
            print(f"[ERROR] {package} 미설치 - 자동 설치 중...")
            try:
                subprocess.check_call([sys.executable, '-m', 'pip', 'install', package])
                print(f"[OK] {package} 설치 완료")
            except subprocess.CalledProcessError as e:
                print(f"[ERROR] {package} 설치 실패: {e}")
                print(f"수동 설치 필요: pip install {package}")


async def collect_partner_info_after_verification(page, business_number, processor):
    """사업자번호 검증 완료 후 거래처 정보 수집 및 저장"""
    try:
        print("      [COLLECT] 거래처 정보 수집 중...")
        await page.wait_for_timeout(1000)  # 정보 로딩 대기
        
        # 거래처 정보 수집
        partner_info = {}
        
        # 1. 상호명 수집
        try:
            company_name = await page.locator("#mf_txppWframe_edtDmnrTnmNmTop").input_value()
            partner_info['company_name'] = company_name.strip() if company_name else ""
            print(f"         상호: {partner_info['company_name']}")
        except Exception as e:
            print(f"         상호 수집 실패: {e}")
            partner_info['company_name'] = ""
        
        # 2. 대표자명 수집
        try:
            representative_name = await page.locator("#mf_txppWframe_edtDmnrRprsFnmTop").input_value()
            partner_info['representative_name'] = representative_name.strip() if representative_name else ""
            print(f"         대표자: {partner_info['representative_name']}")
        except Exception as e:
            print(f"         대표자명 수집 실패: {e}")
            partner_info['representative_name'] = ""
        
        print(f"      [OK] 거래처 정보 수집 완료: {partner_info}")
        return partner_info
        
    except Exception as e:
        print(f"      [ERROR] 거래처 정보 수집 실패: {e}")
        return None


if __name__ == "__main__":
    print("홈택스 세금계산서 자동화 프로그램")
    print("=" * 50)
    
    # 의존성 확인
    check_dependencies()
    
            
            # 최종 상태 확인
            final_url = page.url
            final_title = await page.title()
            
            print(f"\n=== 최종 로그인 결과 ===")
            print(f"최종 URL: {final_url}")
            print(f"최종 제목: {final_title}")
            
            if login_confirmed:
                print("홈택스 자동 로그인 성공!")
                
                # Alert창 X버튼으로 닫기
                print("\n=== Alert창 X버튼 닫기 ===")
                try:
                    # 정확한 X버튼 클릭
                    close_button = page.locator("#mf_txppWframe_UTXPPABB29_wframe_btnCloseInvtSpec")
                    await close_button.wait_for(state="visible", timeout=5000)
                    await close_button.click()
                    print("   X버튼으로 알림창 닫기 완료")
                    await page.wait_for_timeout(2000)
                    
                except Exception as e:
                    print(f"X버튼 클릭 실패: {e}")
                    # 대체 방법으로 Alert 처리
                    try:
                        await page.evaluate("""
                            if (window.confirm) window.confirm = function() { return true; };
                            if (window.alert) window.alert = function() { return true; };
                        """)
                        print("   JavaScript Alert 무력화 완료")
                    except:
                        pass
                
                # 추가 메뉴 네비게이션
                print("\n=== 메뉴 네비게이션 시작 ===")
                await page.wait_for_timeout(3000)  # 더 긴 안정화 대기
                
                try:
                    # 1단계: #mf_wfHeader_wq_uuid_333 선택 (Alert창 닫은 후 첫 번째 메뉴)
                    print("1단계: 첫 번째 메뉴 선택 (#mf_wfHeader_wq_uuid_333)...")
                    
                    first_menu_selectors = [
                        "#mf_wfHeader_wq_uuid_333",
                        "*[id*='wq_uuid_333']",
                        "*[id*='wfHeader'] *[id*='333']",
                        "a[href*='333'], button[id*='333']"
                    ]
                    
                    first_clicked = False
                    for selector in first_menu_selectors:
                        try:
                            print(f"   시도: {selector}")
                            first_menu = page.locator(selector).first
                            await first_menu.wait_for(state="visible", timeout=3000)
                            await first_menu.click()
                            print(f"   첫 번째 메뉴 클릭 성공: {selector}")
                            first_clicked = True
                            break
                        except:
                            continue
                    
                    if not first_clicked:
                        print("   첫 번째 메뉴를 찾을 수 없습니다 - 수동으로 선택하세요")
                        await page.wait_for_timeout(10000)  # 수동 선택 대기
                    else:
                        await page.wait_for_timeout(2000)
                    
                    # 2단계: #combineMenuAtag_4601010100 > span 선택 (두 번째 메뉴)
                    print("2단계: 두 번째 메뉴 선택 (#combineMenuAtag_4601010100)...")
                    
                    second_menu_selectors = [
                        "#combineMenuAtag_4601010100 > span",
                        "#combineMenuAtag_4601010100",
                        "*[id*='combineMenu'][id*='4601010100'] > span",
                        "*[id*='combineMenu'][id*='4601010100']",
                        "a[href*='4601010100'] > span",
                        "a[href*='4601010100']"
                    ]
                    
                    second_clicked = False
                    for selector in second_menu_selectors:
                        try:
                            print(f"   시도: {selector}")
                            second_menu = page.locator(selector).first
                            await second_menu.wait_for(state="visible", timeout=3000)
                            await second_menu.click()
                            print(f"   두 번째 메뉴 클릭 성공: {selector}")
                            second_clicked = True
                            break
                        except:
                            continue
                    
                    if not second_clicked:
                        print("   두 번째 메뉴를 찾을 수 없습니다 - 수동으로 선택하세요")
                        await page.wait_for_timeout(10000)  # 수동 선택 대기
                    else:
                        await page.wait_for_timeout(2000)
                    
                    # 2단계 메뉴 클릭 후 팝업 처리
                    print("3단계: 2단계 메뉴 클릭 후 팝업 처리...")
                    await page.wait_for_timeout(2000)  # 팝업이 나타날 시간 대기
                    
                    try:
                        # Alert 대화상자 자동 처리
                        alert_count = 0
                        def handle_second_dialog(dialog):
                            nonlocal alert_count
                            alert_count += 1
                            print(f"   Alert {alert_count} 감지 및 처리: '{dialog.message}'")
                            asyncio.create_task(dialog.accept())
                        
                        page.on("dialog", handle_second_dialog)
                        
                        # 새 팝업창 확인 및 닫기
                        popup_processed = False
                        for check in range(5):  # 5초간 확인
                            await page.wait_for_timeout(1000)
                            
                            # 새로운 팝업창 확인
                            try:
                                context_pages = page.context.pages
                                if len(context_pages) > 1:
                                    print(f"   새 팝업창 감지: {len(context_pages) - 1}개")
                                    
                                    # 메인 페이지가 아닌 모든 창 닫기
                                    for popup_page in context_pages:
                                        if popup_page != page:
                                            try:
                                                await popup_page.close()
                                                print("   새 팝업창 닫기 완료")
                                                popup_processed = True
                                            except:
                                                pass
                            except:
                                pass
                            
                            # Alert 처리됨 확인
                            if alert_count > 0:
                                print(f"   Alert {alert_count}개 처리 완료")
                                popup_processed = True
                        
                        # 알림창 확인 버튼으로 닫기
                        try:
                            print("   알림창 확인 버튼 찾는 중...")
                            notification_confirm = page.locator("#mf_txppWframe_UTEETZZD02_wframe_btnProcess")
                            await notification_confirm.wait_for(state="visible", timeout=3000)
                            await notification_confirm.click()
                            print("   알림창 확인 버튼 클릭 완료")
                            popup_processed = True
                            await page.wait_for_timeout(1000)
                        except Exception as e:
                            print(f"   알림창 확인 버튼 없음 또는 클릭 실패: {e}")
                        
                        # Alert 핸들러 제거
                        page.remove_listener("dialog", handle_second_dialog)
                        
                        if popup_processed:
                            print("   팝업/Alert 처리 완료")
                        else:
                            print("   팝업/Alert 없음 - 정상 진행")
                            
                    except Exception as popup_error:
                        print(f"   팝업 처리 중 오류: {popup_error}")
                    
                    await page.wait_for_timeout(2000)
                    print("전체 메뉴 네비게이션 완료!")
                    
                except Exception as nav_error:
                    print(f"[ERROR] 메뉴 네비게이션 오류: {nav_error}")
                    print("   수동으로 메뉴를 선택해주세요.")
                
            else:
                print("[WARN]  로그인 상태 확인 필요")
                print("   브라우저에서 직접 확인해주세요.")
            
            print(f"\n=== 세금계산서 자동 처리 시작 ===")
            
            # 선택된 엑셀 데이터를 이용한 세금계산서 자동화
            print("선택된 거래명세표 데이터로 세금계산서 자동화를 시작합니다.")
            print("1. 등록번호 검증 및 자동 입력")
            print("2. 공급일자 자동 입력")  
            print("3. 에러 처리 (번호 error, 미등록 구분)")
            
            await process_tax_invoices_with_selected_data(page, processor)
            
            print(f"\n=== 처리 완료 - 결과 확인 시간 (15초) ===")
            await page.wait_for_timeout(15000)
            
        except Exception as e:
            print(f"오류: {e}")
        finally:
            await browser.close()

def check_dependencies():
    """필수 패키지 확인 및 설치"""
    required_packages = ['openpyxl', 'psutil', 'xlwings', 'pywin32']
    print("[INFO] 의존성 패키지 확인 중...")
    
    for package in required_packages:
        try:
            __import__(package.replace('pywin32', 'win32gui'))  # pywin32는 win32gui로 import
            print(f"[OK] {package} 설치됨")
        except ImportError:
            print(f"[ERROR] {package} 미설치 - 자동 설치 중...")
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", package])
                print(f"[OK] {package} 설치 완료")
            except subprocess.CalledProcessError as e:
                print(f"[ERROR] {package} 설치 실패: {e}")
                print(f"수동 설치 필요: pip install {package}")

async def collect_partner_info_after_verification(page, business_number, processor):
    """사업자번호 검증 완료 후 거래처 정보 수집 및 저장"""
    try:
        print("      [COLLECT] 거래처 정보 수집 중...")
        await page.wait_for_timeout(1000)  # 정보 로딩 대기
        
        # 거래처 정보 수집
        partner_info = {}
        
        # 1. 상호명 수집
        try:
            company_name = await page.locator("#mf_txppWframe_edtDmnrTnmNmTop").input_value()
            partner_info['company_name'] = company_name.strip() if company_name else ""
            print(f"         상호: {partner_info['company_name']}")
        except Exception as e:
            print(f"         상호 수집 실패: {e}")
            partner_info['company_name'] = ""
        
        # 2. 대표자명 수집
        try:
            representative_name = await page.locator("#mf_txppWframe_edtDmnrRprsFnmTop").input_value()
            partner_info['representative_name'] = representative_name.strip() if representative_name else ""
            print(f"         대표자: {partner_info['representative_name']}")
        except Exception as e:
            print(f"         대표자명 수집 실패: {e}")
            partner_info['representative_name'] = ""
        
        # 3. 이메일 앞자리 수집
        try:
            email_front = await page.locator("#mf_txppWframe_edtDmnrMchrgEmlIdTop").input_value()
            partner_info['email_front'] = email_front.strip() if email_front else ""
            print(f"         이메일 앞자리: {partner_info['email_front']}")
        except Exception as e:
            print(f"         이메일 앞자리 수집 실패: {e}")
            partner_info['email_front'] = ""
        
        # 4. 이메일 뒷자리 수집
        try:
            email_back = await page.locator("#mf_txppWframe_edtDmnrMchrgEmlDmanTop").input_value()
            partner_info['email_back'] = email_back.strip() if email_back else ""
            print(f"         이메일 뒷자리: {partner_info['email_back']}")
        except Exception as e:
            print(f"         이메일 뒷자리 수집 실패: {e}")
            partner_info['email_back'] = ""
        
        # 5. 전체 이메일 조합
        if partner_info['email_front'] and partner_info['email_back']:
            partner_info['full_email'] = f"{partner_info['email_front']}@{partner_info['email_back']}"
        else:
            partner_info['full_email'] = ""
        
        # 6. 사업자번호 포함
        partner_info['business_number'] = business_number
        
        # 7. processor에 거래처 정보 저장 (세금계산서 시트 기록용)
        if hasattr(processor, 'partner_info_cache'):
            processor.partner_info_cache[business_number] = partner_info
        else:
            processor.partner_info_cache = {business_number: partner_info}
        
        print(f"         [OK] 거래처 정보 수집 완료: {partner_info['company_name']}")
        
    except Exception as e:
        print(f"      [ERROR] 거래처 정보 수집 실패: {e}")

if __name__ == "__main__":
    print("홈택스 세금계산서 자동화 프로그램")
    print("=" * 50)
    
    # 의존성 확인
    check_dependencies()
    
    # 메인 프로그램 실행
    asyncio.run(hometax_quick_login())
