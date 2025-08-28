# -*- coding: utf-8 -*-
"""
HomeTax 전자세금계산서 자동화 프로그램
1. 엑셀 파일 열기/확인 및 작업할 대상 행 선택
2. 등록번호 검증 및 자동 입력
3. 공급일자 자동 입력
"""

import asyncio
import os
import sys
import subprocess
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from dotenv import load_dotenv
from playwright.async_api import async_playwright
import pandas as pd
from pathlib import Path
import re
from datetime import datetime

def check_and_install_dependencies():
    """필수 의존성 패키지 확인 및 자동 설치"""
    required_packages = {
        'xlwings': 'xlwings>=0.30.0',
        'openpyxl': 'openpyxl>=3.1.0'
    }
    
    missing_packages = []
    
    for package_name, package_spec in required_packages.items():
        try:
            __import__(package_name)
            print(f"✅ {package_name} 설치됨")
        except ImportError:
            missing_packages.append(package_spec)
            print(f"❌ {package_name} 미설치")
    
    if missing_packages:
        print(f"\n📦 {len(missing_packages)}개의 패키지를 설치합니다...")
        for package in missing_packages:
            try:
                print(f"설치 중: {package}")
                subprocess.check_call([sys.executable, "-m", "pip", "install", package])
                print(f"✅ {package} 설치 완료")
            except subprocess.CalledProcessError as e:
                print(f"❌ {package} 설치 실패: {e}")
                print(f"수동으로 설치하세요: pip install {package}")
        print("📦 패키지 설치가 완료되었습니다.\n")
    else:
        print("✅ 모든 필수 패키지가 설치되어 있습니다.\n")

class TaxInvoiceProcessor:
    def __init__(self):
        self.selected_rows = None
        self.selected_data = None
        self.excel_file_path = None
        self.headers = None
        self.processed_data = []
        self.field_mapping = {}
    
    def write_error_to_excel(self, row_number, error_message="error"):
        """엑셀 파일의 지정된 행 발행일 열에 에러 메시지 작성"""
        if not self.excel_file_path:
            print("❌ 엑셀 파일 경로가 없습니다.")
            return False
        
        try:
            from openpyxl import load_workbook
            
            print(f"엑셀 파일에 에러 기록 중: 행 {row_number}, 메시지: {error_message}")
            
            workbook = load_workbook(self.excel_file_path)
            worksheet = workbook.active
            
            # 발행일 열 찾기 (보통 첫 번째 열)
            worksheet.cell(row=row_number, column=1, value=error_message)
            
            workbook.save(self.excel_file_path)
            workbook.close()
            
            print(f"✅ 엑셀 파일에 에러 기록 완료: 행 {row_number}")
            return True
            
        except Exception as e:
            print(f"❌ 엑셀 파일 에러 기록 실패: {e}")
            return False
    
    def select_excel_file(self):
        """엑셀 파일 선택"""
        root = tk.Tk()
        root.withdraw()  # 루트 창 숨기기
        
        file_path = filedialog.askopenfilename(
            title="엑셀 파일을 선택하세요",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if file_path:
            self.excel_file_path = file_path
            print(f"✅ 선택된 엑셀 파일: {file_path}")
            return True
        else:
            print("❌ 파일이 선택되지 않았습니다.")
            return False
    
    def load_excel_data(self):
        """엑셀 데이터 로드"""
        try:
            df = pd.read_excel(self.excel_file_path)
            
            # 헤더 저장
            self.headers = df.columns.tolist()
            
            print(f"✅ 엑셀 데이터 로드 완료")
            print(f"   총 {len(df)}개의 행")
            print(f"   컬럼: {self.headers}")
            
            return df
            
        except Exception as e:
            print(f"❌ 엑셀 파일 로드 실패: {e}")
            return None
    
    def show_row_selection_gui(self, df):
        """행 선택 GUI"""
        root = tk.Tk()
        root.title("전자세금계산서 처리할 행 선택")
        root.geometry("800x600")
        
        # 프레임 생성
        frame = ttk.Frame(root, padding="10")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 트리뷰 생성
        tree = ttk.Treeview(frame, columns=self.headers, show="tree headings", height=15)
        
        # 컬럼 설정
        tree.column("#0", width=50)
        tree.heading("#0", text="행")
        
        for col in self.headers:
            tree.column(col, width=100)
            tree.heading(col, text=str(col)[:15])  # 헤더 길이 제한
        
        # 데이터 삽입
        for index, row in df.iterrows():
            values = [str(val)[:20] if pd.notna(val) else "" for val in row.values]  # 값 길이 제한
            tree.insert("", "end", text=str(index+2), values=values)  # +2는 엑셀 행 번호 (헤더 포함)
        
        # 스크롤바
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # 선택 정보 표시
        info_label = ttk.Label(frame, text="Ctrl+클릭으로 여러 행 선택 가능")
        info_label.grid(row=1, column=0, pady=10)
        
        # 버튼 프레임
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=2, column=0, pady=10)
        
        selected_rows = []
        
        def on_confirm():
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("선택 오류", "처리할 행을 선택해주세요.")
                return
            
            nonlocal selected_rows
            selected_rows = []
            
            for item in selection:
                row_text = tree.item(item)["text"]
                selected_rows.append(int(row_text))
            
            print(f"✅ 선택된 행: {selected_rows}")
            root.quit()
            root.destroy()
        
        def on_cancel():
            root.quit()
            root.destroy()
        
        ttk.Button(button_frame, text="확인", command=on_confirm).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="취소", command=on_cancel).pack(side=tk.LEFT, padx=5)
        
        root.mainloop()
        
        if selected_rows:
            # 선택된 행의 데이터 추출
            self.selected_rows = selected_rows
            selected_data = []
            
            for row_num in selected_rows:
                row_data = df.iloc[row_num-2].to_dict()  # -2는 엑셀 행 번호를 pandas 인덱스로 변환
                row_data['excel_row'] = row_num
                selected_data.append(row_data)
            
            self.selected_data = selected_data
            print(f"✅ {len(selected_data)}개 행의 데이터 추출 완료")
            return True
        
        return False

class TaxInvoiceAutomation:
    def __init__(self, processor):
        self.processor = processor
        self.page = None
    
    async def login_to_hometax(self):
        """HomeTax 로그인"""
        load_dotenv()
        cert_password = os.getenv("PW")
        
        if not cert_password:
            print("❌ .env 파일에 PW 변수가 설정되지 않았습니다.")
            return False
        
        async with async_playwright() as p:
            browser = await p.firefox.launch(headless=False, slow_mo=1000)
            
            try:
                self.page = await browser.new_page()
                self.page.set_default_timeout(10000)
                
                print("🌐 HomeTax 페이지 이동...")
                await self.page.goto("https://hometax.go.kr/websquare/websquare.html?w2xPath=/ui/pp/index_pp.xml&menuCd=index3")
                await self.page.wait_for_load_state('domcontentloaded')
                await self.page.wait_for_timeout(3000)
                
                # 로그인 처리 (기존 hometax_quick.py의 로그인 로직 사용)
                print("🔐 인증서 로그인 처리...")
                await self.handle_certificate_login(cert_password)
                
                # 세금계산서 작성 페이지로 이동
                print("📄 전자세금계산서 작성 페이지 이동...")
                await self.navigate_to_tax_invoice_page()
                
                return True
                
            except Exception as e:
                print(f"❌ HomeTax 로그인 실패: {e}")
                return False
    
    async def handle_certificate_login(self, cert_password):
        """인증서 로그인 처리"""
        try:
            # 공동·금융인증서 버튼 클릭 시도
            button_selectors = [
                "#mf_txppWframe_loginboxFrame_anchor22",
                "#anchor22",
                "a:has-text('공동인증서')",
                "a:has-text('공동·금융인증서')"
            ]
            
            login_clicked = False
            for selector in button_selectors:
                try:
                    await self.page.locator(selector).first.click(timeout=2000)
                    print(f"✅ 로그인 버튼 클릭: {selector}")
                    login_clicked = True
                    break
                except:
                    continue
            
            if not login_clicked:
                print("⚠️ 자동 클릭 실패 - 수동으로 '공동·금융인증서' 버튼을 클릭하세요")
                await self.page.wait_for_timeout(10000)
            
            # dscert iframe 대기 및 처리
            for i in range(15):
                try:
                    await self.page.wait_for_selector("#dscert", timeout=1000)
                    dscert_iframe = self.page.frame_locator("#dscert")
                    await dscert_iframe.locator("body").wait_for(timeout=1000)
                    print("✅ 인증서 창 발견!")
                    break
                except:
                    await self.page.wait_for_timeout(1000)
            
            # 인증서 선택 및 비밀번호 입력
            await self.page.wait_for_timeout(2000)
            
            # 비밀번호 입력
            password_input = dscert_iframe.locator("#input_cert_pw").first
            await password_input.wait_for(state="visible", timeout=5000)
            await password_input.fill(cert_password)
            print("✅ 비밀번호 입력 완료")
            
            # 확인 버튼 클릭
            confirm_btn = dscert_iframe.locator("#btn_confirm_iframe > span").first
            await confirm_btn.wait_for(state="visible", timeout=3000)
            await confirm_btn.click()
            print("✅ 확인 버튼 클릭 완료")
            
            # 로그인 완료 대기
            await self.page.wait_for_timeout(5000)
            
        except Exception as e:
            print(f"❌ 인증서 로그인 처리 실패: {e}")
    
    async def navigate_to_tax_invoice_page(self):
        """전자세금계산서 작성 페이지로 이동"""
        try:
            # Alert 창 닫기
            await self.page.wait_for_timeout(3000)
            try:
                close_button = self.page.locator("#mf_txppWframe_UTXPPABB29_wframe_btnCloseInvtSpec")
                await close_button.wait_for(state="visible", timeout=5000)
                await close_button.click()
                print("✅ Alert 창 닫기 완료")
            except:
                print("⚠️ Alert 창 없음 또는 닫기 실패")
            
            await self.page.wait_for_timeout(2000)
            
            # 전자세금계산서 메뉴 클릭 (실제 셀렉터로 수정 필요)
            # 여기서는 기존 메뉴 네비게이션 로직을 사용
            
        except Exception as e:
            print(f"❌ 세금계산서 페이지 이동 실패: {e}")
    
    async def process_tax_invoices(self):
        """전자세금계산서 자동 처리"""
        if not self.processor.selected_data:
            print("❌ 처리할 데이터가 없습니다.")
            return
        
        print(f"📊 총 {len(self.processor.selected_data)}개 행 처리 시작")
        
        for idx, row_data in enumerate(self.processor.selected_data, 1):
            try:
                print(f"\n[{idx}/{len(self.processor.selected_data)}] 행 {row_data['excel_row']} 처리 중...")
                
                # 1단계: 등록번호 검증 및 입력
                business_number = str(row_data.get('등록번호', '')).strip()
                if not business_number:
                    print("❌ 등록번호가 없습니다.")
                    self.processor.write_error_to_excel(row_data['excel_row'], "번호 error")
                    continue
                
                await self.input_business_number(business_number, row_data)
                
                # 2단계: 공급일자 입력
                supply_date = row_data.get('공급일자', '')
                if supply_date:
                    await self.input_supply_date(supply_date)
                
                print(f"✅ 행 {row_data['excel_row']} 처리 완료")
                await self.page.wait_for_timeout(2000)
                
            except Exception as e:
                print(f"❌ 행 {row_data['excel_row']} 처리 실패: {e}")
                self.processor.write_error_to_excel(row_data['excel_row'], "처리 error")
                continue
        
        print("🎉 전체 처리 완료!")
    
    async def input_business_number(self, business_number, row_data):
        """등록번호 입력 및 검증"""
        try:
            # 등록번호 입력
            business_input = self.page.locator("#mf_txppWframe_edtDmnrBsnoTop")
            await business_input.wait_for(state="visible", timeout=5000)
            await business_input.fill(business_number)
            print(f"   등록번호 입력: {business_number}")
            
            # 확인 버튼 클릭
            confirm_btn = self.page.locator("#mf_txppWframe_btnDmnrBsnoCnfrTop")
            await confirm_btn.click()
            await self.page.wait_for_timeout(2000)
            
            # 상호명 확인
            company_name_input = self.page.locator("#mf_txppWframe_edtDmnrTnmNmTop")
            company_name = await company_name_input.get_attribute("value")
            
            if not company_name or company_name.strip() == "":
                # 등록되지 않은 업체
                print(f"   ⚠️ 미등록 업체: {business_number}")
                self.processor.write_error_to_excel(row_data['excel_row'], "미등록")
                return False
            else:
                # 정상 등록된 업체
                print(f"   ✅ 등록된 업체: {company_name}")
                return True
            
        except Exception as e:
            print(f"   ❌ 등록번호 검증 실패: {e}")
            self.processor.write_error_to_excel(row_data['excel_row'], "번호 error")
            return False
    
    async def input_supply_date(self, supply_date):
        """공급일자 입력"""
        try:
            # 날짜 형식 변환 (필요시)
            if isinstance(supply_date, pd.Timestamp):
                supply_date = supply_date.strftime("%Y%m%d")
            elif isinstance(supply_date, str):
                # 문자열 날짜를 YYYYMMDD 형식으로 변환
                supply_date = supply_date.replace("-", "").replace("/", "").replace(".", "")
            
            # 공급일자 입력
            date_input = self.page.locator("#mf_txppWframe_calWrtDtTop_input")
            await date_input.wait_for(state="visible", timeout=3000)
            await date_input.fill(str(supply_date))
            print(f"   공급일자 입력: {supply_date}")
            
        except Exception as e:
            print(f"   ❌ 공급일자 입력 실패: {e}")

async def main():
    """메인 함수"""
    print("🚀 전자세금계산서 자동화 프로그램 시작")
    print("=" * 50)
    
    # 의존성 확인
    check_and_install_dependencies()
    
    # 프로세서 초기화
    processor = TaxInvoiceProcessor()
    
    # 1단계: 엑셀 파일 선택 및 데이터 로드
    print("📂 1단계: 엑셀 파일 선택")
    if not processor.select_excel_file():
        print("❌ 프로그램을 종료합니다.")
        return
    
    df = processor.load_excel_data()
    if df is None:
        print("❌ 프로그램을 종료합니다.")
        return
    
    # 2단계: 행 선택
    print("\n📊 2단계: 처리할 행 선택")
    if not processor.show_row_selection_gui(df):
        print("❌ 행이 선택되지 않았습니다. 프로그램을 종료합니다.")
        return
    
    # 3단계: HomeTax 자동화
    print("\n🔧 3단계: HomeTax 자동 처리")
    automation = TaxInvoiceAutomation(processor)
    
    if await automation.login_to_hometax():
        await automation.process_tax_invoices()
    else:
        print("❌ HomeTax 로그인 실패")

if __name__ == "__main__":
    print("전자세금계산서 자동화 프로그램")
    asyncio.run(main())