# -*- coding: utf-8 -*-
"""
엑셀 파일의 실제 데이터 행 수 확인
"""

from openpyxl import load_workbook
import os

def check_excel_data_rows():
    """엑셀 파일의 실제 데이터 행 수 확인"""
    excel_file = r"C:\Users\man4k\OneDrive\문서\세금계산서.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"❌ 파일을 찾을 수 없습니다: {excel_file}")
        return
    
    try:
        # 엑셀 파일 로드
        wb = load_workbook(excel_file)
        
        # 모든 시트명 출력
        print(f"엑셀 파일 분석: {excel_file}")
        print(f"전체 시트 목록: {wb.sheetnames}")
        
        # "거래처" 시트 찾기
        if "거래처" in wb.sheetnames:
            ws = wb["거래처"]
            print(f"선택된 워크시트: {ws.title}")
        else:
            print("❌ '거래처' 시트를 찾을 수 없습니다.")
            print("사용 가능한 시트:")
            for sheet_name in wb.sheetnames:
                print(f"  - {sheet_name}")
            wb.close()
            return
        
        print(f"openpyxl 최대 행: {ws.max_row}")
        print(f"openpyxl 최대 열: {ws.max_column}")
        print()
        
        # 전체 행 확인
        print("=== 전체 행 데이터 분석 ===")
        data_rows = 0
        header_found = False
        
        for row in range(1, ws.max_row + 1):
            # A열, B열, D열 값 확인
            a_value = ws.cell(row=row, column=1).value or ""
            b_value = ws.cell(row=row, column=2).value or ""
            d_value = ws.cell(row=row, column=4).value or ""
            
            a_str = str(a_value).strip()
            b_str = str(b_value).strip()
            d_str = str(d_value).strip()
            
            # 헤더 행 확인
            if not header_found and ("등록" in b_str or "사업자" in b_str):
                print(f"행 {row}: [헤더] A='{a_str}' B='{b_str}' D='{d_str}'")
                header_found = True
                continue
            
            # 데이터 행 확인 (B열에 값이 있으면 데이터로 간주)
            if b_str and b_str not in ['', 'None', '등록번호', '사업자번호', '사업자등록번호']:
                data_rows += 1
                if data_rows <= 5:  # 처음 5개만 출력
                    print(f"행 {row}: [데이터{data_rows}] A='{a_str}' B='{b_str}' D='{d_str}'")
            elif row <= 10:  # 처음 10행 중 빈 행도 표시
                print(f"행 {row}: [빈행] A='{a_str}' B='{b_str}' D='{d_str}'")
        
        print()
        print("결과:")
        print(f"   총 행 수: {ws.max_row}행")
        print(f"   헤더 행: 1행 (발견됨)" if header_found else "   헤더 행: 미발견")
        print(f"   실제 데이터 행: {data_rows}행")
        print(f"   처리 가능한 행 범위: 2 ~ {ws.max_row}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")

if __name__ == "__main__":
    check_excel_data_rows()